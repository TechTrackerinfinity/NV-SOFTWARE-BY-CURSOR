import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
import shutil
import zipfile
import tempfile
from datetime import datetime, timedelta
import json
import re
import hashlib
import time
import io
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
import base64
from io import BytesIO
from werkzeug.utils import secure_filename
import threading
import logging
import logging.handlers
import traceback
import sys

# Configure logging
def setup_logging():
    """Configure the logging system for the application."""
    log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs')
    os.makedirs(log_dir, exist_ok=True)
    
    # Create a logger
    logger = logging.getLogger('diamond_app')
    logger.setLevel(logging.DEBUG)
    
    # Create handlers
    # Console handler for development
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    
    # File handler for all logs
    file_handler = logging.handlers.RotatingFileHandler(
        os.path.join(log_dir, 'diamond_app.log'),
        maxBytes=10485760,  # 10MB
        backupCount=10
    )
    file_handler.setLevel(logging.DEBUG)
    
    # Error file handler for errors only
    error_file_handler = logging.handlers.RotatingFileHandler(
        os.path.join(log_dir, 'error.log'),
        maxBytes=10485760,  # 10MB
        backupCount=10
    )
    error_file_handler.setLevel(logging.ERROR)
    
    # Create formatters
    console_formatter = logging.Formatter('%(levelname)s - %(message)s')
    file_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    
    # Add formatters to handlers
    console_handler.setFormatter(console_formatter)
    file_handler.setFormatter(file_formatter)
    error_file_handler.setFormatter(file_formatter)
    
    # Add handlers to logger
    logger.addHandler(console_handler)
    logger.addHandler(file_handler)
    logger.addHandler(error_file_handler)
    
    return logger

# Initialize logger
logger = setup_logging()
logger.info("Starting Diamond Accounting Application")

app = Flask(__name__)
app.secret_key = 'diamond_business_secret_key'

# Add global error handlers
@app.errorhandler(404)
def page_not_found(e):
    logger.info(f"404 error: {request.path}")
    return render_template('error.html', 
                          error_code=404, 
                          error_message="The page you're looking for doesn't exist."), 404

@app.errorhandler(500)
def internal_server_error(e):
    logger.error(f"500 error: {str(e)}")
    logger.error(f"Request: {request.path} {request.method}")
    logger.error(f"Form data: {request.form}")
    return render_template('error.html', 
                          error_code=500, 
                          error_message="Something went wrong on our end. Please try again later."), 500

@app.errorhandler(Exception)
def handle_exception(e):
    # Get the exception info
    exc_info = sys.exc_info()
    
    # Format the traceback
    tb_lines = traceback.format_exception(*exc_info)
    tb_text = ''.join(tb_lines)
    
    # Log the error with traceback
    logger.error(f"Unhandled exception: {str(e)}")
    logger.error(f"Request: {request.path} {request.method}")
    logger.error(f"Form data: {request.form}")
    logger.error(f"Traceback: {tb_text}")
    
    # Return a user-friendly error page
    return render_template('error.html', 
                          error_code=500, 
                          error_message="An unexpected error occurred. Our team has been notified."), 500

# Add custom Jinja2 filters
@app.template_filter('format_currency')
def format_currency(value):
    if value is None:
        return "0.00"
    try:
        value = float(value)
        return "{:,.2f}".format(value)
    except (ValueError, TypeError):
        return "0.00"

@app.template_filter('format_datetime')
def format_datetime(timestamp):
    """Format a timestamp into a readable date and time."""
    try:
        dt = datetime.fromtimestamp(timestamp)
        return dt.strftime('%Y-%m-%d %H:%M:%S')
    except (ValueError, TypeError):
        return "Unknown"

@app.template_filter('format_size')
def format_size(size_bytes):
    """Format a file size in bytes to a human-readable format."""
    try:
        size_bytes = float(size_bytes)
        if size_bytes < 1024:
            return f"{size_bytes:.2f} B"
        elif size_bytes < 1024 * 1024:
            return f"{size_bytes / 1024:.2f} KB"
        elif size_bytes < 1024 * 1024 * 1024:
            return f"{size_bytes / (1024 * 1024):.2f} MB"
        else:
            return f"{size_bytes / (1024 * 1024 * 1024):.2f} GB"
    except (ValueError, TypeError):
        return "Unknown"

@app.template_filter('file_stats')
def file_stats(file_path):
    """Get file statistics."""
    try:
        if os.path.exists(file_path):
            stats = os.stat(file_path)
            return {
                'size': stats.st_size,
                'mtime': stats.st_mtime,
                'ctime': stats.st_ctime
            }
        return {'size': 0, 'mtime': 0, 'ctime': 0}
    except Exception:
        return {'size': 0, 'mtime': 0, 'ctime': 0}

# Add min and max functions to Jinja2 environment
app.jinja_env.globals.update(min=min)
app.jinja_env.globals.update(max=max)

# Ensure data directory exists
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
os.makedirs(DATA_DIR, exist_ok=True)

# Define file paths
PURCHASES_FILE = os.path.join(DATA_DIR, 'purchases.xlsx')
SALES_FILE = os.path.join(DATA_DIR, 'sales.xlsx')
PAYMENTS_FILE = os.path.join(DATA_DIR, 'payments.xlsx')
INVENTORY_FILE = os.path.join(DATA_DIR, 'inventory.xlsx')
ROUGH_INVENTORY_FILE = os.path.join(DATA_DIR, 'rough_inventory.xlsx')

# Define backup directory
BACKUP_DIR = os.path.join(DATA_DIR, 'backup')
os.makedirs(BACKUP_DIR, exist_ok=True)

# Function to create a backup of all data files
def create_backup():
    """
    Create a backup of all data files in a zip file.
    Returns the path to the created backup file.
    """
    try:
        logger.info("Creating backup of all data files")
        
        # Create a timestamp for the backup file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_file = os.path.join(BACKUP_DIR, f'diamond_data_backup_{timestamp}.zip')
        logger.debug(f"Backup file path: {backup_file}")
        
        # Check if all required files exist
        missing_files = []
        for file_path in [PURCHASES_FILE, SALES_FILE, PAYMENTS_FILE, INVENTORY_FILE, ROUGH_INVENTORY_FILE]:
            if not os.path.exists(file_path):
                missing_files.append(os.path.basename(file_path))
        
        if missing_files:
            logger.warning(f"The following files are missing and will not be included in the backup: {', '.join(missing_files)}")
        
        # Create a zip file containing all data files
        with zipfile.ZipFile(backup_file, 'w') as zipf:
            for file_path in [PURCHASES_FILE, SALES_FILE, PAYMENTS_FILE, INVENTORY_FILE, ROUGH_INVENTORY_FILE]:
                if os.path.exists(file_path):
                    zipf.write(file_path, os.path.basename(file_path))
                    logger.debug(f"Added file to backup: {os.path.basename(file_path)}")
        
        # Keep only the 10 most recent backups
        backup_files = sorted([os.path.join(BACKUP_DIR, f) for f in os.listdir(BACKUP_DIR) 
                              if f.startswith('diamond_data_backup_') and f.endswith('.zip')],
                             key=os.path.getmtime, reverse=True)
        
        for old_backup in backup_files[10:]:
            try:
                os.remove(old_backup)
                logger.debug(f"Removed old backup: {os.path.basename(old_backup)}")
            except Exception as e:
                logger.warning(f"Could not remove old backup {old_backup}: {str(e)}")
        
        logger.info(f"Backup created successfully: {os.path.basename(backup_file)}")
        return backup_file
    except Exception as e:
        logger.error(f"Error creating backup: {str(e)}")
        # Log the traceback
        exc_info = sys.exc_info()
        tb_lines = traceback.format_exception(*exc_info)
        tb_text = ''.join(tb_lines)
        logger.error(f"Traceback: {tb_text}")
        return None

# Function to restore from a backup file
def restore_from_backup(backup_file):
    """
    Restore data from a backup zip file.
    """
    try:
        logger.info(f"Restoring from backup: {os.path.basename(backup_file)}")
        
        # Validate the backup file
        if not os.path.exists(backup_file):
            logger.error(f"Backup file does not exist: {backup_file}")
            return False
            
        # Verify it's a valid zip file
        try:
            with zipfile.ZipFile(backup_file, 'r') as zipf:
                # Check if the zip file contains the expected files
                file_list = zipf.namelist()
                logger.debug(f"Files in backup: {', '.join(file_list)}")
                
                expected_files = [os.path.basename(f) for f in 
                                 [PURCHASES_FILE, SALES_FILE, PAYMENTS_FILE, INVENTORY_FILE, ROUGH_INVENTORY_FILE]]
                missing_files = [f for f in expected_files if f not in file_list]
                
                if missing_files:
                    logger.warning(f"The following files are missing from the backup: {', '.join(missing_files)}")
                    # Proceed anyway, but warn the user
        except zipfile.BadZipFile:
            logger.error(f"{backup_file} is not a valid zip file")
            return False
            
        # Create backup of current data before restoring
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        pre_restore_backup = os.path.join(BACKUP_DIR, f'pre_restore_backup_{timestamp}.zip')
        try:
            with zipfile.ZipFile(pre_restore_backup, 'w') as zipf:
                for file_path in [PURCHASES_FILE, SALES_FILE, PAYMENTS_FILE, INVENTORY_FILE, ROUGH_INVENTORY_FILE]:
                    if os.path.exists(file_path):
                        zipf.write(file_path, os.path.basename(file_path))
            logger.info(f"Created pre-restore backup: {os.path.basename(pre_restore_backup)}")
        except Exception as e:
            logger.warning(f"Could not create pre-restore backup: {str(e)}")
            # Continue with restore even if pre-restore backup fails
        
        # Create a temporary directory for extraction
        with tempfile.TemporaryDirectory() as temp_dir:
            logger.debug(f"Created temporary directory for extraction: {temp_dir}")
            
            # Extract the backup file
            with zipfile.ZipFile(backup_file, 'r') as zipf:
                zipf.extractall(temp_dir)
                logger.debug(f"Extracted backup to temporary directory")
            
            # Copy the extracted files to the data directory
            for file_name in os.listdir(temp_dir):
                src_path = os.path.join(temp_dir, file_name)
                dst_path = os.path.join(DATA_DIR, file_name)
                shutil.copy2(src_path, dst_path)
                logger.debug(f"Copied {file_name} to data directory")
        
        # Validate the restored data
        try:
            logger.info("Validating restored data")
            fix_data_types()
            validate_data_consistency()
            logger.info("Data validation successful after restore")
        except Exception as e:
            logger.warning(f"Data validation after restore encountered issues: {str(e)}")
            # Continue anyway, the data might still be usable
        
        logger.info("Restore completed successfully")
        return True
    except Exception as e:
        logger.error(f"Error restoring from backup: {str(e)}")
        # Log the traceback
        exc_info = sys.exc_info()
        tb_lines = traceback.format_exception(*exc_info)
        tb_text = ''.join(tb_lines)
        logger.error(f"Traceback: {tb_text}")
        return False

# Function to fix data type inconsistencies in Excel files
def fix_data_types():
    """
    Fix data type inconsistencies in Excel files.
    Ensures that numeric columns are stored as appropriate numeric types.
    """
    try:
        # Fix inventory.xlsx
        if os.path.exists(INVENTORY_FILE):
            inventory_df = pd.read_excel(INVENTORY_FILE)
            if not inventory_df.empty:
                # Convert numeric columns to appropriate types
                numeric_columns = ['carats', 'purchase_price', 'market_value']
                for col in numeric_columns:
                    if col in inventory_df.columns:
                        inventory_df[col] = pd.to_numeric(inventory_df[col], errors='coerce')
                
                # Save the fixed DataFrame
                inventory_df.to_excel(INVENTORY_FILE, index=False)
                print(f"Fixed data types in {INVENTORY_FILE}")
        
        # Fix rough_inventory.xlsx
        if os.path.exists(ROUGH_INVENTORY_FILE):
            rough_inventory_df = pd.read_excel(ROUGH_INVENTORY_FILE)
            if not rough_inventory_df.empty:
                # Convert numeric columns to appropriate types
                numeric_columns = ['weight', 'pieces', 'purchase_price']
                for col in numeric_columns:
                    if col in rough_inventory_df.columns:
                        rough_inventory_df[col] = pd.to_numeric(rough_inventory_df[col], errors='coerce')
                
                # Ensure required columns exist
                required_columns = ['rough_id', 'kapan_no', 'shape_category']
                for col in required_columns:
                    if col not in rough_inventory_df.columns:
                        rough_inventory_df[col] = ''
                
                # Save the fixed DataFrame
                rough_inventory_df.to_excel(ROUGH_INVENTORY_FILE, index=False)
                print(f"Fixed data types in {ROUGH_INVENTORY_FILE}")
        
        # Fix purchases.xlsx
        if os.path.exists(PURCHASES_FILE):
            purchases_df = pd.read_excel(PURCHASES_FILE)
            if not purchases_df.empty:
                # Convert numeric columns to appropriate types
                numeric_columns = ['amount', 'Carat', 'Price Per Carat', 'Total Amount']
                for col in numeric_columns:
                    if col in purchases_df.columns:
                        purchases_df[col] = pd.to_numeric(purchases_df[col], errors='coerce')
                
                # Save the fixed DataFrame
                purchases_df.to_excel(PURCHASES_FILE, index=False)
                print(f"Fixed data types in {PURCHASES_FILE}")
        
        # Fix sales.xlsx
        if os.path.exists(SALES_FILE):
            sales_df = pd.read_excel(SALES_FILE)
            if not sales_df.empty:
                # Convert numeric columns to appropriate types
                numeric_columns = ['Carat', 'Price Per Carat', 'Total Amount', 'carat', 'price_per_carat', 'total_amount_usd']
                for col in numeric_columns:
                    if col in sales_df.columns:
                        sales_df[col] = pd.to_numeric(sales_df[col], errors='coerce')
                
                # Save the fixed DataFrame
                sales_df.to_excel(SALES_FILE, index=False)
                print(f"Fixed data types in {SALES_FILE}")
        
        # Fix payments.xlsx
        if os.path.exists(PAYMENTS_FILE):
            payments_df = pd.read_excel(PAYMENTS_FILE)
            if not payments_df.empty:
                # Convert numeric columns to appropriate types
                numeric_columns = ['total_amount', 'paid_amount', 'pending_amount']
                for col in numeric_columns:
                    if col in payments_df.columns:
                        payments_df[col] = pd.to_numeric(payments_df[col], errors='coerce')
                
                # Save the fixed DataFrame
                payments_df.to_excel(PAYMENTS_FILE, index=False)
                print(f"Fixed data types in {PAYMENTS_FILE}")
        
        return True
    except Exception as e:
        print(f"Error fixing data types: {str(e)}")
        return False

# Function to enhance Excel file formatting
def enhance_excel_formatting(file_path, sheet_name='Sheet1'):
    # Load the workbook
    try:
        from openpyxl import load_workbook, Workbook
        
        # Check if file exists first
        if not os.path.exists(file_path):
            # Create a new workbook if file doesn't exist
            wb = Workbook()
            sheet = wb.active
            sheet.title = sheet_name
            wb.save(file_path)
            print(f"Created new Excel file at {file_path}")
        
        # Load the workbook directly with openpyxl
        wb = load_workbook(file_path)
        
        # Get the active sheet
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            # If the specified sheet doesn't exist, use the active sheet
            sheet = wb.active
            # Only set title if it's not already set
            if sheet is not None:
                sheet.title = sheet_name
        
        # Ensure sheet is not None before proceeding
        if sheet is None:
            raise ValueError(f"Could not find or create sheet '{sheet_name}' in workbook")
        
        # Format headers
        if sheet.max_row > 0:  # Only proceed if there are rows in the sheet
            for col_num, column_title in enumerate(sheet[1], 1):
                if column_title is None or column_title.value is None:
                    continue
                
                cell = sheet.cell(row=1, column=col_num)
                cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                
                # Adjust column width based on content
                column_letter = get_column_letter(col_num)
                max_length = 0
                for cell in sheet[column_letter]:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = min(adjusted_width, 30)
        
            # Format data cells
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    cell.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
                    
                    # Align numeric values to the right
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.number_format = '#,##0.00'
                    else:
                        cell.alignment = Alignment(vertical='center')
        
            # Add conditional formatting for payment status
            status_column = None
            for col_num, column_title in enumerate(sheet[1], 1):
                if column_title.value == 'Payment Status':
                    status_column = get_column_letter(col_num)
                    break
            
            if status_column:
                # Pending - Light Red
                pending_rule = CellIsRule(operator='equal', formula=['"Pending"'], 
                                        stopIfTrue=True, fill=PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'))
                sheet.conditional_formatting.add(f'{status_column}2:{status_column}{sheet.max_row}', pending_rule)
                
                # Completed - Light Green
                completed_rule = CellIsRule(operator='equal', formula=['"Completed"'], 
                                        stopIfTrue=True, fill=PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid'))
                sheet.conditional_formatting.add(f'{status_column}2:{status_column}{sheet.max_row}', completed_rule)
                
                # Partial - Light Yellow
                partial_rule = CellIsRule(operator='equal', formula=['"Partial"'], 
                                        stopIfTrue=True, fill=PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'))
                sheet.conditional_formatting.add(f'{status_column}2:{status_column}{sheet.max_row}', partial_rule)
            
            # Format amount columns
            for col_num, column_title in enumerate(sheet[1], 1):
                if column_title.value and 'Amount' in str(column_title.value):
                    column_letter = get_column_letter(col_num)
                    for row in range(2, sheet.max_row + 1):
                        cell = sheet[f'{column_letter}{row}']
                        cell.number_format = '#,##0.00'
                        
                        # Add currency symbol based on column name
                        if 'USD' in str(column_title.value):
                            cell.number_format = '"$"#,##0.00'
                        elif 'INR' in str(column_title.value):
                            cell.number_format = '"₹"#,##0.00'
            
            # Format date columns
            for col_num, column_title in enumerate(sheet[1], 1):
                if column_title.value and 'Date' in str(column_title.value):
                    column_letter = get_column_letter(col_num)
                    for row in range(2, sheet.max_row + 1):
                        cell = sheet[f'{column_letter}{row}']
                        if cell.value:
                            cell.number_format = 'yyyy-mm-dd'
            
            # Add alternating row colors for better readability
            for row in range(2, sheet.max_row + 1):
                if row % 2 == 0:  # Even rows
                    for col in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row, column=col)
                        if not cell.fill.start_color.index == 'FFCCCC' and not cell.fill.start_color.index == 'CCFFCC' and not cell.fill.start_color.index == 'FFFFCC':
                            cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
            
            # Freeze the header row
            sheet.freeze_panes = 'A2'
            
            # Add data validation for Payment Status column if it exists
            if status_column:
                from openpyxl.worksheet.datavalidation import DataValidation
                dv = DataValidation(type="list", formula1='"Pending,Completed,Partial"', allow_blank=True)
                sheet.add_data_validation(dv)
                dv.add(f'{status_column}2:{status_column}{sheet.max_row}')
        
        # Save the workbook
        wb.save(file_path)
        print(f"Successfully enhanced Excel formatting for {file_path}")
        
        # Create summary sheet if there's enough data
        if sheet.max_row > 1:
            try:
                create_summary_sheet(file_path)
                create_dashboard(file_path)
            except Exception as e:
                print(f"Error creating summary or dashboard: {str(e)}")
        
        return file_path
    except Exception as e:
        print(f"Error in enhance_excel_formatting: {str(e)}")
        # Return the original file path even if formatting failed
        return file_path

# Function to create a summary sheet with charts and key metrics
def create_summary_sheet(file_path):
    try:
        from openpyxl import load_workbook
        from openpyxl.chart import PieChart, BarChart, Reference
        from openpyxl.chart.series import DataPoint
        
        # Load the workbook
        wb = load_workbook(file_path)
        
        # Ensure there's at least one sheet in the workbook
        if len(wb.sheetnames) == 0:
            print(f"Error: No sheets found in workbook {file_path}")
            return
            
        data_sheet = wb.active
        
        # Check if summary sheet already exists, if not create it
        if 'Summary' in wb.sheetnames:
            wb.remove(wb['Summary'])
        
        summary_sheet = wb.create_sheet(title='Summary')
        
        # Set column widths
        for col in range(1, 10):
            summary_sheet.column_dimensions[get_column_letter(col)].width = 15
        
        # Add title
        summary_sheet['A1'] = 'Diamond Transaction Summary'
        summary_sheet['A1'].font = Font(name='Arial', size=16, bold=True)
        summary_sheet.merge_cells('A1:I1')
        summary_sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Add section headers
        summary_sheet['A3'] = 'Key Metrics'
        summary_sheet['A3'].font = Font(name='Arial', size=14, bold=True)
        summary_sheet['A3'].fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        summary_sheet['A3'].font = Font(color='FFFFFF', bold=True)
        summary_sheet.merge_cells('A3:I3')
        summary_sheet['A3'].alignment = Alignment(horizontal='center')
        
        # Calculate key metrics
        total_records = max(0, data_sheet.max_row - 1)  # Ensure it's not negative
        
        # Find column indices
        carat_col = None
        amount_usd_col = None
        amount_inr_col = None
        status_col = None
        date_col = None
        
        # Check if data_sheet has any rows before trying to access them
        if data_sheet.max_row > 0:
            for col_num, cell in enumerate(data_sheet[1], 1):
                if cell.value == 'Carat':
                    carat_col = col_num
                elif cell.value == 'Total Amount USD':
                    amount_usd_col = col_num
                elif cell.value == 'Total Amount INR':
                    amount_inr_col = col_num
                elif cell.value == 'Payment Status':
                    status_col = col_num
                elif cell.value == 'Date':
                    date_col = col_num
        
        # Calculate totals
        total_carat = 0
        total_amount_usd = 0
        total_amount_inr = 0
        pending_count = 0
        completed_count = 0
        partial_count = 0
        
        for row in range(2, data_sheet.max_row + 1):
            if carat_col:
                carat_value = data_sheet.cell(row=row, column=carat_col).value
                if carat_value and isinstance(carat_value, (int, float)):
                    total_carat += carat_value
            
            if amount_usd_col:
                amount_value = data_sheet.cell(row=row, column=amount_usd_col).value
                if amount_value and isinstance(amount_value, (int, float)):
                    total_amount_usd += amount_value
            
            if amount_inr_col:
                amount_value = data_sheet.cell(row=row, column=amount_inr_col).value
                if amount_value and isinstance(amount_value, (int, float)):
                    total_amount_inr += amount_value
            
            if status_col:
                status_value = data_sheet.cell(row=row, column=status_col).value
                if status_value == 'Pending':
                    pending_count += 1
                elif status_value == 'Completed':
                    completed_count += 1
                elif status_value == 'Partial':
                    partial_count += 1
        
        # Add metrics to summary sheet
        summary_sheet['A5'] = 'Total Records:'
        summary_sheet['B5'] = total_records
        summary_sheet['A5'].font = Font(bold=True)
        
        summary_sheet['A6'] = 'Total Carat:'
        summary_sheet['B6'] = total_carat
        summary_sheet['B6'].number_format = '#,##0.00'
        summary_sheet['A6'].font = Font(bold=True)
        
        summary_sheet['A7'] = 'Total Amount (USD):'
        summary_sheet['B7'] = total_amount_usd
        summary_sheet['B7'].number_format = '"$"#,##0.00'
        summary_sheet['A7'].font = Font(bold=True)
        
        summary_sheet['A8'] = 'Total Amount (INR):'
        summary_sheet['B8'] = total_amount_inr
        summary_sheet['B8'].number_format = '"₹"#,##0.00'
        summary_sheet['A8'].font = Font(bold=True)
        
        summary_sheet['A9'] = 'Pending Payments:'
        summary_sheet['B9'] = pending_count
        summary_sheet['A9'].font = Font(bold=True)
        
        summary_sheet['A10'] = 'Completed Payments:'
        summary_sheet['B10'] = completed_count
        summary_sheet['A10'].font = Font(bold=True)
        
        summary_sheet['A11'] = 'Partial Payments:'
        summary_sheet['B11'] = partial_count
        summary_sheet['A11'].font = Font(bold=True)
        
        # Add borders to metrics
        for row in range(5, 12):
            for col in range(1, 3):
                cell = summary_sheet.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Add payment status chart
        if status_col and (pending_count > 0 or completed_count > 0 or partial_count > 0):
            summary_sheet['A13'] = 'Payment Status Distribution'
            summary_sheet['A13'].font = Font(name='Arial', size=14, bold=True)
            summary_sheet['A13'].fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            summary_sheet['A13'].font = Font(color='FFFFFF', bold=True)
            summary_sheet.merge_cells('A13:I13')
            summary_sheet['A13'].alignment = Alignment(horizontal='center')
            
            # Create data for pie chart
            summary_sheet['A15'] = 'Status'
            summary_sheet['B15'] = 'Count'
            summary_sheet['A15'].font = Font(bold=True)
            summary_sheet['B15'].font = Font(bold=True)
            
            summary_sheet['A16'] = 'Pending'
            summary_sheet['B16'] = pending_count
            
            summary_sheet['A17'] = 'Completed'
            summary_sheet['B17'] = completed_count
            
            summary_sheet['A18'] = 'Partial'
            summary_sheet['B18'] = partial_count
            
            # Create pie chart
            pie = PieChart()
            labels = Reference(summary_sheet, min_col=1, min_row=16, max_row=18)
            data = Reference(summary_sheet, min_col=2, min_row=15, max_row=18)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Payment Status Distribution"
            
            # Add custom colors to pie chart slices
            slice1 = DataPoint(idx=0, fill=PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'))
            slice2 = DataPoint(idx=1, fill=PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid'))
            slice3 = DataPoint(idx=2, fill=PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'))
            pie.series[0].data_points = [slice1, slice2, slice3]
            
            pie.height = 10
            pie.width = 10
            summary_sheet.add_chart(pie, "D15")
        
        # Add monthly trend if date column exists
        if date_col and total_records > 0:
            summary_sheet['A25'] = 'Monthly Transaction Trend'
            summary_sheet['A25'].font = Font(name='Arial', size=14, bold=True)
            summary_sheet['A25'].fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            summary_sheet['A25'].font = Font(color='FFFFFF', bold=True)
            summary_sheet.merge_cells('A25:I25')
            summary_sheet['A25'].alignment = Alignment(horizontal='center')
            
            # Collect monthly data
            monthly_data = {}
            
            for row in range(2, data_sheet.max_row + 1):
                date_value = data_sheet.cell(row=row, column=date_col).value
                if date_value:
                    try:
                        if isinstance(date_value, str):
                            date_value = datetime.strptime(date_value, '%Y-%m-%d')
                        
                        month_key = f"{date_value.year}-{date_value.month:02d}"
                        
                        if month_key not in monthly_data:
                            monthly_data[month_key] = {
                                'count': 0,
                                'amount_usd': 0,
                                'amount_inr': 0
                            }
                        
                        monthly_data[month_key]['count'] += 1
                        
                        if amount_usd_col:
                            amount_value = data_sheet.cell(row=row, column=amount_usd_col).value
                            if amount_value and isinstance(amount_value, (int, float)):
                                monthly_data[month_key]['amount_usd'] += amount_value
                        
                        if amount_inr_col:
                            amount_value = data_sheet.cell(row=row, column=amount_inr_col).value
                            if amount_value and isinstance(amount_value, (int, float)):
                                monthly_data[month_key]['amount_inr'] += amount_value
                    except Exception as e:
                        print(f"Error processing date in row {row}: {str(e)}")
                        continue
            
            # Sort months chronologically
            sorted_months = sorted(monthly_data.keys())
            
            # Add data for bar chart
            summary_sheet['A27'] = 'Month'
            summary_sheet['B27'] = 'Count'
            summary_sheet['C27'] = 'Amount (USD)'
            summary_sheet['D27'] = 'Amount (INR)'
            
            summary_sheet['A27'].font = Font(bold=True)
            summary_sheet['B27'].font = Font(bold=True)
            summary_sheet['C27'].font = Font(bold=True)
            summary_sheet['D27'].font = Font(bold=True)
            
            for i, month in enumerate(sorted_months, 28):
                summary_sheet[f'A{i}'] = month
                summary_sheet[f'B{i}'] = monthly_data[month]['count']
                summary_sheet[f'C{i}'] = monthly_data[month]['amount_usd']
                summary_sheet[f'D{i}'] = monthly_data[month]['amount_inr']
                
                summary_sheet[f'C{i}'].number_format = '"$"#,##0.00'
                summary_sheet[f'D{i}'].number_format = '"₹"#,##0.00'
            
            # Create bar chart for transaction count
            if len(sorted_months) > 0:
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = "Monthly Transaction Count"
                chart.y_axis.title = "Number of Transactions"
                chart.x_axis.title = "Month"
                
                data = Reference(summary_sheet, min_col=2, min_row=27, max_row=27+len(sorted_months))
                cats = Reference(summary_sheet, min_col=1, min_row=28, max_row=27+len(sorted_months))
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.shape = 4
                chart.height = 10
                chart.width = 15
                
                summary_sheet.add_chart(chart, "F27")
        
        # Apply styling to the entire summary sheet
        for row in summary_sheet.iter_rows():
            for cell in row:
                if not cell.font.bold:
                    cell.font = Font(name='Arial')
        
        # Save the workbook
        wb.save(file_path)
    except Exception as e:
        print(f"Error in create_summary_sheet: {str(e)}")
        # Don't re-raise the exception, just log it and continue

# Function to create a dashboard with visual elements
def create_dashboard(file_path):
    try:
        from openpyxl import load_workbook
        from openpyxl.chart import PieChart, BarChart, LineChart, Reference
        from openpyxl.chart.series import DataPoint
        from openpyxl.drawing.image import Image
        
        # Load the workbook
        wb = load_workbook(file_path)
        
        # Ensure there's at least one sheet in the workbook
        if len(wb.sheetnames) == 0:
            print(f"Error: No sheets found in workbook {file_path}")
            return
            
        data_sheet = wb.active
        
        # Check if dashboard sheet already exists, if not create it
        if 'Dashboard' in wb.sheetnames:
            wb.remove(wb['Dashboard'])
        
        dashboard = wb.create_sheet(title='Dashboard', index=0)  # Make it the first sheet
        
        # Set column widths
        for col in range(1, 15):
            dashboard.column_dimensions[get_column_letter(col)].width = 15
        
        # Add title
        dashboard['A1'] = 'DIAMOND BUSINESS DASHBOARD'
        dashboard['A1'].font = Font(name='Arial', size=20, bold=True, color='4F81BD')
        dashboard.merge_cells('A1:N1')
        dashboard['A1'].alignment = Alignment(horizontal='center')
        
        # Add subtitle with current date
        current_date = datetime.now().strftime('%Y-%m-%d')
        dashboard['A2'] = f'Generated on: {current_date}'
        dashboard['A2'].font = Font(name='Arial', size=10, italic=True)
        dashboard.merge_cells('A2:N2')
        dashboard['A2'].alignment = Alignment(horizontal='center')
        
        # Find column indices
        carat_col = None
        amount_usd_col = None
        amount_inr_col = None
        status_col = None
        date_col = None
        
        # Check if data_sheet has any rows before trying to access them
        if data_sheet.max_row > 0:
            for col_num, cell in enumerate(data_sheet[1], 1):
                if cell.value == 'Carat':
                    carat_col = col_num
                elif cell.value == 'Total Amount USD':
                    amount_usd_col = col_num
                elif cell.value == 'Total Amount INR':
                    amount_inr_col = col_num
                elif cell.value == 'Payment Status':
                    status_col = col_num
                elif cell.value == 'Date':
                    date_col = col_num
        
        # Calculate key metrics
        total_records = max(0, data_sheet.max_row - 1)  # Ensure it's not negative
        total_carat = 0
        total_amount_usd = 0
        total_amount_inr = 0
        pending_count = 0
        completed_count = 0
        partial_count = 0
        
        # Data for monthly trends
        monthly_data = {}
        
        for row in range(2, data_sheet.max_row + 1):
            if carat_col:
                carat_value = data_sheet.cell(row=row, column=carat_col).value
                if carat_value and isinstance(carat_value, (int, float)):
                    total_carat += carat_value
            
            if amount_usd_col:
                amount_value = data_sheet.cell(row=row, column=amount_usd_col).value
                if amount_value and isinstance(amount_value, (int, float)):
                    total_amount_usd += amount_value
            
            if amount_inr_col:
                amount_value = data_sheet.cell(row=row, column=amount_inr_col).value
                if amount_value and isinstance(amount_value, (int, float)):
                    total_amount_inr += amount_value
            
            if status_col:
                status_value = data_sheet.cell(row=row, column=status_col).value
                if status_value == 'Pending':
                    pending_count += 1
                elif status_value == 'Completed':
                    completed_count += 1
                elif status_value == 'Partial':
                    partial_count += 1
            
            # Collect monthly data
            if date_col:
                date_value = data_sheet.cell(row=row, column=date_col).value
                if date_value:
                    try:
                        if isinstance(date_value, str):
                            date_value = datetime.strptime(date_value, '%Y-%m-%d')
                        
                        month_key = f"{date_value.year}-{date_value.month:02d}"
                        month_name = f"{calendar.month_name[date_value.month]} {date_value.year}"
                        
                        if month_key not in monthly_data:
                            monthly_data[month_key] = {
                                'name': month_name,
                                'count': 0,
                                'amount_usd': 0,
                                'amount_inr': 0,
                                'carat': 0
                            }
                        
                        monthly_data[month_key]['count'] += 1
                        
                        if amount_usd_col:
                            amount_value = data_sheet.cell(row=row, column=amount_usd_col).value
                            if amount_value and isinstance(amount_value, (int, float)):
                                monthly_data[month_key]['amount_usd'] += amount_value
                        
                        if amount_inr_col:
                            amount_value = data_sheet.cell(row=row, column=amount_inr_col).value
                            if amount_value and isinstance(amount_value, (int, float)):
                                monthly_data[month_key]['amount_inr'] += amount_value
                        
                        if carat_col:
                            carat_value = data_sheet.cell(row=row, column=carat_col).value
                            if carat_value and isinstance(carat_value, (int, float)):
                                monthly_data[month_key]['carat'] += carat_value
                    except Exception as e:
                        print(f"Error processing date in row {row}: {str(e)}")
                        continue
        
        # Sort months chronologically
        sorted_months = sorted(monthly_data.keys())
        
        # Create KPI cards
        dashboard['A4'] = 'KEY PERFORMANCE INDICATORS'
        dashboard['A4'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        dashboard['A4'].fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        dashboard.merge_cells('A4:N4')
        dashboard['A4'].alignment = Alignment(horizontal='center')
        
        # KPI Card 1: Total Records
        dashboard['B6'] = 'Total Records'
        dashboard['B6'].font = Font(name='Arial', size=12, bold=True)
        dashboard['B6'].alignment = Alignment(horizontal='center')
        dashboard.merge_cells('B6:D6')
        
        dashboard['B7'] = total_records
        dashboard['B7'].font = Font(name='Arial', size=24, bold=True, color='4F81BD')
        dashboard['B7'].alignment = Alignment(horizontal='center')
        dashboard.merge_cells('B7:D7')
        
        # Add border to KPI card
        for row in range(6, 8):
            for col in range(2, 5):
                cell = dashboard.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # KPI Card 2: Total Carat
        dashboard['F6'] = 'Total Carat'
        dashboard['F6'].font = Font(name='Arial', size=12, bold=True)
        dashboard['F6'].alignment = Alignment(horizontal='center')
        dashboard.merge_cells('F6:H6')
        
        dashboard['F7'] = total_carat
        dashboard['F7'].font = Font(name='Arial', size=24, bold=True, color='4F81BD')
        dashboard['F7'].alignment = Alignment(horizontal='center')
        dashboard['F7'].number_format = '#,##0.00'
        dashboard.merge_cells('F7:H7')
        
        # Add border to KPI card
        for row in range(6, 8):
            for col in range(6, 9):
                cell = dashboard.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # KPI Card 3: Total Amount USD
        dashboard['J6'] = 'Total Amount (USD)'
        dashboard['J6'].font = Font(name='Arial', size=12, bold=True)
        dashboard['J6'].alignment = Alignment(horizontal='center')
        dashboard.merge_cells('J6:L6')
        
        dashboard['J7'] = total_amount_usd
        dashboard['J7'].font = Font(name='Arial', size=24, bold=True, color='4F81BD')
        dashboard['J7'].alignment = Alignment(horizontal='center')
        dashboard['J7'].number_format = '"$"#,##0.00'
        dashboard.merge_cells('J7:L7')
        
        # Add border to KPI card
        for row in range(6, 8):
            for col in range(10, 13):
                cell = dashboard.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # KPI Card 4: Total Amount INR
        dashboard['B9'] = 'Total Amount (INR)'
        dashboard['B9'].font = Font(name='Arial', size=12, bold=True)
        dashboard['B9'].alignment = Alignment(horizontal='center')
        dashboard.merge_cells('B9:D9')
        
        dashboard['B10'] = total_amount_inr
        dashboard['B10'].font = Font(name='Arial', size=24, bold=True, color='4F81BD')
        dashboard['B10'].alignment = Alignment(horizontal='center')
        dashboard['B10'].number_format = '"₹"#,##0.00'
        dashboard.merge_cells('B10:D10')
        
        # Add border to KPI card
        for row in range(9, 11):
            for col in range(2, 5):
                cell = dashboard.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # KPI Card 5: Pending Payments
        dashboard['F9'] = 'Pending Payments'
        dashboard['F9'].font = Font(name='Arial', size=12, bold=True)
        dashboard['F9'].alignment = Alignment(horizontal='center')
        dashboard.merge_cells('F9:H9')
        
        dashboard['F10'] = pending_count
        dashboard['F10'].font = Font(name='Arial', size=24, bold=True, color='FF0000')
        dashboard['F10'].alignment = Alignment(horizontal='center')
        dashboard.merge_cells('F10:H10')
        
        # Add border to KPI card
        for row in range(9, 11):
            for col in range(6, 9):
                cell = dashboard.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # KPI Card 6: Completed Payments
        dashboard['J9'] = 'Completed Payments'
        dashboard['J9'].font = Font(name='Arial', size=12, bold=True)
        dashboard['J9'].alignment = Alignment(horizontal='center')
        dashboard.merge_cells('J9:L9')
        
        dashboard['J10'] = completed_count
        dashboard['J10'].font = Font(name='Arial', size=24, bold=True, color='008000')
        dashboard['J10'].alignment = Alignment(horizontal='center')
        dashboard.merge_cells('J10:L10')
        
        # Add border to KPI card
        for row in range(9, 11):
            for col in range(10, 13):
                cell = dashboard.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Add Payment Status Chart
        if status_col and (pending_count > 0 or completed_count > 0 or partial_count > 0):
            dashboard['A12'] = 'PAYMENT STATUS DISTRIBUTION'
            dashboard['A12'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
            dashboard['A12'].fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            dashboard.merge_cells('A12:N12')
            dashboard['A12'].alignment = Alignment(horizontal='center')
            
            # Create data for pie chart
            dashboard['B14'] = 'Status'
            dashboard['C14'] = 'Count'
            dashboard['B14'].font = Font(bold=True)
            dashboard['C14'].font = Font(bold=True)
            
            dashboard['B15'] = 'Pending'
            dashboard['C15'] = pending_count
            dashboard['B15'].fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            
            dashboard['B16'] = 'Completed'
            dashboard['C16'] = completed_count
            dashboard['B16'].fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
            
            dashboard['B17'] = 'Partial'
            dashboard['C17'] = partial_count
            dashboard['B17'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            
            # Create pie chart
            pie = PieChart()
            labels = Reference(dashboard, min_col=2, min_row=15, max_row=17)
            data = Reference(dashboard, min_col=3, min_row=14, max_row=17)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Payment Status Distribution"
            
            # Add custom colors to pie chart slices
            slice1 = DataPoint(idx=0, fill=PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'))
            slice2 = DataPoint(idx=1, fill=PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid'))
            slice3 = DataPoint(idx=2, fill=PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'))
            pie.series[0].data_points = [slice1, slice2, slice3]
            
            pie.height = 10
            pie.width = 10
            dashboard.add_chart(pie, "E14")
        
        # Add Monthly Trend Charts
        if date_col and len(sorted_months) > 0:
            dashboard['A22'] = 'MONTHLY TRENDS'
            dashboard['A22'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
            dashboard['A22'].fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            dashboard.merge_cells('A22:N22')
            dashboard['A22'].alignment = Alignment(horizontal='center')
            
            # Create data for charts
            dashboard['B24'] = 'Month'
            dashboard['C24'] = 'Transactions'
            dashboard['D24'] = 'Amount (USD)'
            dashboard['E24'] = 'Amount (INR)'
            dashboard['F24'] = 'Carat'
            
            dashboard['B24'].font = Font(bold=True)
            dashboard['C24'].font = Font(bold=True)
            dashboard['D24'].font = Font(bold=True)
            dashboard['E24'].font = Font(bold=True)
            dashboard['F24'].font = Font(bold=True)
            
            for i, month_key in enumerate(sorted_months, 25):
                month_data = monthly_data[month_key]
                dashboard[f'B{i}'] = month_data['name']
                dashboard[f'C{i}'] = month_data['count']
                dashboard[f'D{i}'] = month_data['amount_usd']
                dashboard[f'E{i}'] = month_data['amount_inr']
                dashboard[f'F{i}'] = month_data['carat']
                
                dashboard[f'D{i}'].number_format = '"$"#,##0.00'
                dashboard[f'E{i}'].number_format = '"₹"#,##0.00'
                dashboard[f'F{i}'].number_format = '#,##0.00'
            
            # Create line chart for monthly transactions
            line = LineChart()
            line.title = "Monthly Transaction Count"
            line.style = 12
            line.y_axis.title = "Number of Transactions"
            line.x_axis.title = "Month"
            
            data = Reference(dashboard, min_col=3, min_row=24, max_row=24+len(sorted_months))
            cats = Reference(dashboard, min_col=2, min_row=25, max_row=24+len(sorted_months))
            line.add_data(data, titles_from_data=True)
            line.set_categories(cats)
            
            line.height = 10
            line.width = 15
            dashboard.add_chart(line, "H24")
            
            # Create line chart for monthly carat
            line2 = LineChart()
            line2.title = "Monthly Carat Volume"
            line2.style = 13
            line2.y_axis.title = "Total Carat"
            line2.x_axis.title = "Month"
            
            data = Reference(dashboard, min_col=6, min_row=24, max_row=24+len(sorted_months))
            cats = Reference(dashboard, min_col=2, min_row=25, max_row=24+len(sorted_months))
            line2.add_data(data, titles_from_data=True)
            line2.set_categories(cats)
            
            line2.height = 10
            line2.width = 15
            dashboard.add_chart(line2, "H38")
        
        # Add footer
        footer_row = dashboard.max_row + 5
        dashboard[f'A{footer_row}'] = 'Generated by Diamond Accounting App'
        dashboard[f'A{footer_row}'].font = Font(name='Arial', size=10, italic=True)
        dashboard.merge_cells(f'A{footer_row}:N{footer_row}')
        dashboard[f'A{footer_row}'].alignment = Alignment(horizontal='center')
        
        # Save the workbook
        wb.save(file_path)
    except Exception as e:
        print(f"Error in create_dashboard: {str(e)}")
        # Don't re-raise the exception, just log it and continue

# Initialize Excel files if they don't exist
def initialize_excel_files():
    """Initialize Excel files if they don't exist."""
    os.makedirs('data', exist_ok=True)
    
    # Initialize purchases file
    if not os.path.exists(PURCHASES_FILE):
        df = pd.DataFrame(columns=['Date', 'Party', 'Diamond Type', 'Carats', 'Rate per Carat', 
                                  'Total Amount USD', 'Exchange Rate', 'Total Amount INR', 'Payment Status', 'Notes'])
        df.to_excel(PURCHASES_FILE, index=False)
        enhance_excel_formatting(PURCHASES_FILE)
    
    # Initialize sales file
    if not os.path.exists(SALES_FILE):
        df = pd.DataFrame(columns=['Date', 'Party', 'Diamond Type', 'Carats', 'Rate per Carat', 
                                  'Total Amount USD', 'Exchange Rate', 'Total Amount INR', 'Payment Status', 'Notes'])
        df.to_excel(SALES_FILE, index=False)
        enhance_excel_formatting(SALES_FILE)
    
    # Initialize payments file
    if not os.path.exists(PAYMENTS_FILE):
        df = pd.DataFrame(columns=['id', 'type', 'name', 'total_amount', 'paid_amount', 'pending_amount', 
                                 'status', 'payment_date', 'payment_method', 'notes'])
        df.to_excel(PAYMENTS_FILE, index=False)
    
    # Initialize inventory file
    if not os.path.exists(INVENTORY_FILE):
        df = pd.DataFrame(columns=['id', 'description', 'shape', 'carats', 'color', 'clarity', 'cut', 
                                  'purchase_price', 'market_value', 'status', 'location', 'purchase_date', 'notes'])
        df.to_excel(INVENTORY_FILE, index=False)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/test')
def test():
    return "Hello, World! The application is working."

@app.route('/buy', methods=['GET', 'POST'])
def buy():
    if request.method == 'POST':
        try:
            # Get the diamond type from the form
            diamond_type = request.form.get('diamond_type', 'polished')
            
            # Get common form data
            date = request.form.get('date') or None
            party = request.form.get('party') or None
            notes = request.form.get('notes') or None
            
            if diamond_type == 'polished':
                # Process polished diamond purchase
                stone_id = request.form.get('stone_id') or None
                platform = request.form.get('platform') or None
                
                # Validate numeric fields for polished diamond
                try:
                    carat = float(request.form.get('carat_detail', 0))
                    price_per_carat = float(request.form.get('price_per_carat_inr', 0))
                    
                    if carat <= 0:
                        flash('Carat must be greater than 0', 'danger')
                        return redirect(url_for('buy'))
                    
                    if price_per_carat <= 0:
                        flash('Price per carat must be greater than 0', 'danger')
                        return redirect(url_for('buy'))
                        
                    # Calculate total price
                    total_price = carat * price_per_carat
                    
                    # Add to polished inventory
                    if os.path.exists(INVENTORY_FILE):
                        inventory_df = pd.read_excel(INVENTORY_FILE)
                    else:
                        inventory_df = pd.DataFrame(columns=['id', 'description', 'shape', 'carats', 'color', 'clarity', 'cut', 
                                                          'purchase_price', 'market_value', 'status', 'location', 'purchase_date', 
                                                          'notes', 'rough_id'])
                    
                    # Generate a unique ID
                    item_id = f"D{int(time.time())}"
                    
                    # Get shape from form
                    shape = get_shape_from_form(request.form)
                    
                    # Create new item
                    new_item = {
                        'id': item_id,
                        'description': request.form.get('description') or f"{shape} {carat}ct Diamond",
                        'shape': shape,
                        'carats': carat,
                        'color': get_color_from_form(request.form),
                        'clarity': request.form.get('clarity_detail') or '',
                        'cut': request.form.get('cut_detail') or '',
                        'purchase_price': total_price,
                        'market_value': total_price * 1.2,  # 20% markup as default
                        'status': 'In Stock',
                        'location': request.form.get('location') or '',
                        'purchase_date': date,
                        'notes': notes,
                        'rough_id': ''
                    }
                    
                    # Add to DataFrame
                    inventory_df = pd.concat([inventory_df, pd.DataFrame([new_item])], ignore_index=True)
                    
                    # Save to Excel
                    inventory_df.to_excel(INVENTORY_FILE, index=False)
                    
                    # Record purchase in purchases.xlsx
                    if os.path.exists(os.path.join(DATA_DIR, 'purchases.xlsx')):
                        purchases_df = pd.read_excel(os.path.join(DATA_DIR, 'purchases.xlsx'))
                    else:
                        purchases_df = pd.DataFrame(columns=['id', 'date', 'party', 'item_id', 'description', 'amount', 'payment_status', 'notes', 'diamond_type'])
                    
                    # Create purchase record
                    purchase_id = f"P{int(time.time())}"
                    purchase_record = {
                        'id': purchase_id,
                        'date': date,
                        'party': party,
                        'item_id': item_id,
                        'description': new_item['description'],
                        'amount': total_price,
                        'payment_status': request.form.get('payment_status') or 'Pending',
                        'notes': notes,
                        'diamond_type': 'polished'
                    }
                    
                    purchases_df = pd.concat([purchases_df, pd.DataFrame([purchase_record])], ignore_index=True)
                    purchases_df.to_excel(os.path.join(DATA_DIR, 'purchases.xlsx'), index=False)
                    
                    flash(f'Polished diamond purchase recorded successfully! Added to inventory with ID: {item_id}', 'success')
                    
                except (ValueError, TypeError) as e:
                    flash(f'Invalid numeric value: {str(e)}', 'danger')
                    return redirect(url_for('buy'))
                
            elif diamond_type == 'rough':
                # Process rough diamond purchase
                rough_id = request.form.get('rough_id') or None
                kapan_no = request.form.get('kapan_no') or None
                source = request.form.get('source') or None
                origin = request.form.get('origin') or None
                
                # Validate numeric fields for rough diamond
                try:
                    weight = float(request.form.get('weight', 0))
                    pieces = int(request.form.get('pieces', 1))
                    purchase_price = float(request.form.get('purchase_price', 0))
                    
                    if weight <= 0:
                        flash('Weight must be greater than 0', 'danger')
                        return redirect(url_for('buy'))
                    
                    if purchase_price <= 0:
                        flash('Purchase price must be greater than 0', 'danger')
                        return redirect(url_for('buy'))
                    
                    # Add to rough inventory
                    if os.path.exists(ROUGH_INVENTORY_FILE):
                        rough_inventory_df = pd.read_excel(ROUGH_INVENTORY_FILE)
                    else:
                        rough_inventory_df = pd.DataFrame(columns=['id', 'lot_id', 'description', 'source', 'origin', 
                                                                'weight', 'pieces', 'purchase_price', 'purchase_date', 
                                                                'status', 'location', 'notes', 'image_path', 'rough_id', 'kapan_no',
                                                                'shape_category'])
                    
                    # Generate a unique ID and lot ID
                    item_id = f"R{int(time.time())}"
                    lot_id = f"LOT-{str(int(time.time()))[-4:]}" if pieces > 1 else ""
                    
                    # Create new item
                    new_item = {
                        'id': item_id,
                        'lot_id': lot_id,
                        'description': request.form.get('description') or f"Rough Diamond {weight}ct",
                        'source': source,
                        'origin': origin,
                        'weight': weight,
                        'pieces': pieces,
                        'purchase_price': purchase_price,
                        'purchase_date': date,
                        'status': 'In Stock',
                        'location': request.form.get('location') or '',
                        'notes': notes,
                        'image_path': '',
                        'rough_id': rough_id,
                        'kapan_no': kapan_no,
                        'shape_category': get_shape_from_form(request.form, 'rough')
                    }
                    
                    # Add to DataFrame
                    rough_inventory_df = pd.concat([rough_inventory_df, pd.DataFrame([new_item])], ignore_index=True)
                    
                    # Save to Excel
                    rough_inventory_df.to_excel(ROUGH_INVENTORY_FILE, index=False)
                    
                    # Record purchase in purchases.xlsx
                    if os.path.exists(os.path.join(DATA_DIR, 'purchases.xlsx')):
                        purchases_df = pd.read_excel(os.path.join(DATA_DIR, 'purchases.xlsx'))
                    else:
                        purchases_df = pd.DataFrame(columns=['id', 'date', 'party', 'item_id', 'description', 'amount', 'payment_status', 'notes', 'diamond_type'])
                    
                    # Create purchase record
                    purchase_id = f"P{int(time.time())}"
                    purchase_record = {
                        'id': purchase_id,
                        'date': date,
                        'party': party,
                        'item_id': item_id,
                        'description': new_item['description'],
                        'amount': purchase_price,
                        'payment_status': request.form.get('payment_status') or 'Pending',
                        'notes': notes,
                        'diamond_type': 'rough'
                    }
                    
                    purchases_df = pd.concat([purchases_df, pd.DataFrame([purchase_record])], ignore_index=True)
                    purchases_df.to_excel(os.path.join(DATA_DIR, 'purchases.xlsx'), index=False)
                    
                    flash('Rough diamond purchase recorded successfully!', 'success')
                    
                except (ValueError, TypeError) as e:
                    flash(f'Invalid numeric value: {str(e)}', 'danger')
                    return redirect(url_for('buy'))
            
            return redirect(url_for('buy'))
            
        except Exception as e:
            flash(f'Error processing purchase: {str(e)}', 'danger')
            return redirect(url_for('buy'))
    
    # For GET requests, display the buy form
    return render_template('buy.html', today_date=datetime.now().strftime('%Y-%m-%d'))

@app.route('/add_payment', methods=['POST'])
def add_payment():
    try:
        # Get form data
        payment_type = request.form.get('type')
        name = request.form.get('name')
        amount = float(request.form.get('amount'))
        payment_date = request.form.get('payment_date')
        payment_method = request.form.get('payment_method')
        notes = request.form.get('notes')

        # Load existing payments
        if os.path.exists(PAYMENTS_FILE):
            df = pd.read_excel(PAYMENTS_FILE)
        else:
            df = pd.DataFrame(columns=['id', 'type', 'name', 'total_amount', 'paid_amount', 'pending_amount', 
                                     'status', 'payment_date', 'payment_method', 'notes'])

        # Generate new payment ID
        new_id = str(len(df) + 1)

        # Create new payment record
        new_payment = {
            'id': new_id,
            'type': payment_type,
            'name': name,
            'total_amount': amount,
            'paid_amount': 0,
            'pending_amount': amount,
            'status': 'pending',
            'payment_date': pd.to_datetime(payment_date),
            'payment_method': payment_method,
            'notes': notes
        }

        # Append new payment to DataFrame
        df = pd.concat([df, pd.DataFrame([new_payment])], ignore_index=True)

        # Save to Excel
        df.to_excel(PAYMENTS_FILE, index=False)

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/payment_details/<payment_id>')
def payment_details(payment_id):
    try:
        if not os.path.exists(PAYMENTS_FILE):
            flash('No payment records found', 'error')
            return redirect(url_for('payments'))

        df = pd.read_excel(PAYMENTS_FILE)
        payment = df[df['id'] == payment_id].iloc[0].to_dict()

        return render_template('payment_details.html', payment=payment)
    except Exception as e:
        flash(f'Error loading payment details: {str(e)}', 'error')
        return redirect(url_for('payments'))

@app.route('/edit_payment/<payment_id>', methods=['GET', 'POST'])
def edit_payment(payment_id):
    try:
        if not os.path.exists(PAYMENTS_FILE):
            flash('No payment records found', 'error')
            return redirect(url_for('payments'))

        df = pd.read_excel(PAYMENTS_FILE)
        payment = df[df['id'] == payment_id].iloc[0].to_dict()

        if request.method == 'POST':
            # Update payment details
            df.loc[df['id'] == payment_id, 'name'] = request.form.get('name')
            df.loc[df['id'] == payment_id, 'payment_date'] = pd.to_datetime(request.form.get('payment_date'))
            df.loc[df['id'] == payment_id, 'payment_method'] = request.form.get('payment_method')
            df.loc[df['id'] == payment_id, 'notes'] = request.form.get('notes')

            # Update payment status based on amounts
            total_amount = float(df.loc[df['id'] == payment_id, 'total_amount'].iloc[0])
            paid_amount = float(request.form.get('paid_amount', 0))
            pending_amount = total_amount - paid_amount

            df.loc[df['id'] == payment_id, 'paid_amount'] = paid_amount
            df.loc[df['id'] == payment_id, 'pending_amount'] = pending_amount

            if pending_amount <= 0:
                status = 'completed'
            elif paid_amount > 0:
                status = 'partial'
            else:
                status = 'pending'

            df.loc[df['id'] == payment_id, 'status'] = status

            # Save changes
            df.to_excel(PAYMENTS_FILE, index=False)
            flash('Payment updated successfully', 'success')
            return redirect(url_for('payments'))

        return render_template('edit_payment.html', payment=payment)
    except Exception as e:
        flash(f'Error editing payment: {str(e)}', 'error')
        return redirect(url_for('payments'))

def get_inventory_status_color(status):
    """
    Returns a Bootstrap color class based on inventory status.
    """
    status = str(status).lower()
    if status == 'in stock':
        return 'success'
    elif status == 'sold':
        return 'danger'
    elif status == 'reserved':
        return 'warning'
    elif status == 'processing':
        return 'info'
    else:
        return 'secondary'

def get_color_from_form(form_data):
    """
    Extract the color value from form data based on color type.
    """
    color_type = form_data.get('color_type', 'White')
    
    if color_type == 'White':
        return form_data.get('white_color', '')
    elif color_type == 'Fancy':
        fancy_color = form_data.get('fancy_color', '')
        fancy_intensity = form_data.get('fancy_intensity', '')
        
        # Check if custom fancy color is provided
        custom_fancy_color = form_data.get('custom_fancy_color', '')
        if custom_fancy_color and custom_fancy_color.strip():
            fancy_color = custom_fancy_color
        
        # Combine intensity and color
        if fancy_intensity and fancy_color:
            return f"{fancy_intensity} {fancy_color}"
        else:
            return fancy_color
    
    return ''

def get_shape_from_form(form_data, form_type='polished'):
    """
    Extract the shape value from form data, handling custom shapes.
    """
    if form_type == 'polished':
        shape = form_data.get('shape_select', '')
        custom_shape = form_data.get('custom_shape', '')
    else:  # rough
        shape = form_data.get('shape_category', '')
        custom_shape = form_data.get('rough_custom_shape', '')
    
    # Check if custom shape is provided
    if shape == 'Other' and custom_shape and custom_shape.strip():
        return custom_shape
    
    return shape

@app.route('/inventory')
def inventory():
    try:
        # Load inventory data
        if os.path.exists(INVENTORY_FILE):
            inventory_df = pd.read_excel(INVENTORY_FILE)
        else:
            inventory_df = pd.DataFrame(columns=['id', 'description', 'shape', 'carats', 'color', 'clarity', 'cut', 
                                              'purchase_price', 'market_value', 'status', 'location', 'purchase_date', 
                                              'notes', 'rough_id'])
            inventory_df.to_excel(INVENTORY_FILE, index=False)
        
        # Apply filters if provided
        shape = request.args.get('shape')
        status = request.args.get('status')
        min_carats = request.args.get('min_carats')
        max_carats = request.args.get('max_carats')
        min_price = request.args.get('min_price')
        max_price = request.args.get('max_price')
        rough_id = request.args.get('rough_id')
        
        filtered_df = inventory_df.copy()
        
        if shape:
            filtered_df = filtered_df[filtered_df['shape'] == shape]
        if status:
            filtered_df = filtered_df[filtered_df['status'] == status]
        if min_carats:
            filtered_df = filtered_df[filtered_df['carats'] >= float(min_carats)]
        if max_carats:
            filtered_df = filtered_df[filtered_df['carats'] <= float(max_carats)]
        if min_price:
            filtered_df = filtered_df[filtered_df['market_value'] >= float(min_price)]
        if max_price:
            filtered_df = filtered_df[filtered_df['market_value'] <= float(max_price)]
        if rough_id:
            filtered_df = filtered_df[filtered_df['rough_id'] == rough_id]
        
        # Calculate totals
        total_items = len(filtered_df)
        total_value = filtered_df['market_value'].sum() if not filtered_df.empty else 0
        total_carats = filtered_df['carats'].sum() if not filtered_df.empty else 0
        
        # Count low stock items (items with only 1 in stock)
        low_stock_count = len(filtered_df[filtered_df['status'].str.lower() == 'in stock']) if not filtered_df.empty else 0
        
        # Load rough inventory for dropdown
        rough_inventory_options = []
        if os.path.exists(ROUGH_INVENTORY_FILE):
            rough_inventory_df = pd.read_excel(ROUGH_INVENTORY_FILE)
            for _, row in rough_inventory_df.iterrows():
                if row.get('status') != 'Processed' and row.get('status') != 'Sold':
                    rough_inventory_options.append({
                        'id': row.get('id', ''),
                        'description': row.get('description', ''),
                        'weight': row.get('weight', 0),
                        'pieces': row.get('pieces', 1)
                    })
        
        # Prepare inventory items for display
        inventory = []
        for _, row in filtered_df.iterrows():
            # Get rough stone info if available
            rough_info = ""
            if 'rough_id' in row and row['rough_id']:
                if os.path.exists(ROUGH_INVENTORY_FILE):
                    rough_inventory_df = pd.read_excel(ROUGH_INVENTORY_FILE)
                    rough_data = rough_inventory_df[rough_inventory_df['id'] == row['rough_id']]
                    if not rough_data.empty:
                        rough_info = f"{rough_data.iloc[0].get('description', '')} ({rough_data.iloc[0].get('weight', 0)} ct)"
            
            item = {
                'id': row.get('id', ''),
                'description': row.get('description', ''),
                'shape': row.get('shape', ''),
                'carats': row.get('carats', 0),
                'color': row.get('color', ''),
                'clarity': row.get('clarity', ''),
                'cut': row.get('cut', ''),
                'purchase_price': row.get('purchase_price', 0),
                'market_value': row.get('market_value', 0),
                'status': row.get('status', ''),
                'status_color': get_inventory_status_color(str(row.get('status', ''))),
                'location': row.get('location', ''),
                'notes': row.get('notes', ''),
                'rough_id': row.get('rough_id', ''),
                'rough_info': rough_info
            }
            inventory.append(item)
        
        # Prepare chart data
        shape_counts = filtered_df['shape'].value_counts() if not filtered_df.empty else pd.Series()
        shape_labels = shape_counts.index.tolist()
        shape_data = shape_counts.values.tolist()
        
        clarity_counts = filtered_df['clarity'].value_counts() if not filtered_df.empty else pd.Series()
        clarity_labels = clarity_counts.index.tolist()
        clarity_data = clarity_counts.values.tolist()
        
        return render_template('inventory.html', 
                             inventory=inventory,
                             total_items=total_items,
                             total_value=total_value,
                             total_carats=total_carats,
                             low_stock_count=low_stock_count,
                             shape_labels=shape_labels,
                             shape_data=shape_data,
                             clarity_labels=clarity_labels,
                             clarity_data=clarity_data,
                             rough_inventory_options=rough_inventory_options)
    except Exception as e:
        flash(f'Error loading inventory: {str(e)}', 'error')
        return render_template('inventory.html', 
                             inventory=[],
                             total_items=0,
                             total_value=0,
                             total_carats=0,
                             low_stock_count=0,
                             shape_labels=[],
                             shape_data=[],
                             clarity_labels=[],
                             clarity_data=[],
                             rough_inventory_options=[])

@app.route('/add_inventory_item', methods=['POST'])
def add_inventory_item():
    try:
        # Get form data
        description = request.form.get('description')
        shape = request.form.get('shape')
        carats = float(request.form.get('carats'))
        color = request.form.get('color')
        clarity = request.form.get('clarity')
        cut = request.form.get('cut')
        purchase_price = float(request.form.get('purchase_price'))
        market_value = float(request.form.get('market_value'))
        status = request.form.get('status')
        location = request.form.get('location')
        notes = request.form.get('notes')
        rough_id = request.form.get('rough_id')
        
        # Load existing inventory
        if os.path.exists(INVENTORY_FILE):
            inventory_df = pd.read_excel(INVENTORY_FILE)
        else:
            inventory_df = pd.DataFrame(columns=['id', 'description', 'shape', 'carats', 'color', 'clarity', 'cut', 
                                              'purchase_price', 'market_value', 'status', 'location', 'purchase_date', 
                                              'notes', 'rough_id'])
        
        # Generate a unique ID
        item_id = f"D{int(time.time())}"
        
        # Create new item
        new_item = {
            'id': item_id,
            'description': description,
            'shape': shape,
            'carats': carats,
            'color': color,
            'clarity': clarity,
            'cut': cut,
            'purchase_price': purchase_price,
            'market_value': market_value,
            'status': status,
            'location': location,
            'purchase_date': pd.Timestamp.now().strftime('%Y-%m-%d'),
            'notes': notes,
            'rough_id': rough_id
        }
        
        # Add to DataFrame
        inventory_df = pd.concat([inventory_df, pd.DataFrame([new_item])], ignore_index=True)
        
        # Save to Excel
        inventory_df.to_excel(INVENTORY_FILE, index=False)
        
        # Update rough stone status if a rough stone was selected
        if rough_id:
            if os.path.exists(ROUGH_INVENTORY_FILE):
                rough_inventory_df = pd.read_excel(ROUGH_INVENTORY_FILE)
                rough_item_index = rough_inventory_df[rough_inventory_df['id'] == rough_id].index
                
                if len(rough_item_index) > 0:
                    # Check if all carats from rough stone have been processed
                    rough_weight = rough_inventory_df.loc[rough_item_index[0], 'weight']
                    
                    # Get all polished diamonds from this rough stone
                    polished_diamonds = inventory_df[inventory_df['rough_id'] == rough_id]
                    total_polished_carats = polished_diamonds['carats'].sum()
                    
                    # If more than 90% of the rough weight has been processed, mark as processed
                    if total_polished_carats >= (rough_weight * 0.9):
                        rough_inventory_df.loc[rough_item_index[0], 'status'] = 'Processed'
                        rough_inventory_df.to_excel(ROUGH_INVENTORY_FILE, index=False)
                        flash('Rough stone marked as processed as most of its weight has been converted to polished diamonds.', 'info')
        
        flash('Inventory item added successfully!', 'success')
        return redirect(url_for('inventory'))
    except Exception as e:
        flash(f'Error adding inventory item: {str(e)}', 'error')
        return redirect(url_for('inventory'))

@app.route('/inventory_item_details/<item_id>')
def inventory_item_details(item_id):
    try:
        # Load inventory data
        if not os.path.exists(INVENTORY_FILE):
            flash('Inventory file not found', 'error')
            return redirect(url_for('inventory'))
        
        inventory_df = pd.read_excel(INVENTORY_FILE)
        
        # Find the item
        item_data = inventory_df[inventory_df['id'] == item_id]
        
        if item_data.empty:
            flash('Item not found', 'error')
            return redirect(url_for('inventory'))
        
        # Get rough stone info if available
        rough_stone = None
        if 'rough_id' in item_data.columns and item_data.iloc[0]['rough_id']:
            rough_id = item_data.iloc[0]['rough_id']
            if os.path.exists(ROUGH_INVENTORY_FILE):
                rough_inventory_df = pd.read_excel(ROUGH_INVENTORY_FILE)
                rough_data = rough_inventory_df[rough_inventory_df['id'] == rough_id]
                if not rough_data.empty:
                    rough_stone = {
                        'id': rough_data.iloc[0].get('id', ''),
                        'description': rough_data.iloc[0].get('description', ''),
                        'source': rough_data.iloc[0].get('source', ''),
                        'origin': rough_data.iloc[0].get('origin', ''),
                        'weight': rough_data.iloc[0].get('weight', 0),
                        'pieces': rough_data.iloc[0].get('pieces', 1),
                        'status': rough_data.iloc[0].get('status', ''),
                        'status_color': get_inventory_status_color(str(rough_data.iloc[0].get('status', ''))),
                        'image_path': rough_data.iloc[0].get('image_path', '')
                    }
        
        # Prepare item for display
        item = {
            'id': item_data.iloc[0].get('id', ''),
            'description': item_data.iloc[0].get('description', ''),
            'shape': item_data.iloc[0].get('shape', ''),
            'carats': item_data.iloc[0].get('carats', 0),
            'color': item_data.iloc[0].get('color', ''),
            'clarity': item_data.iloc[0].get('clarity', ''),
            'cut': item_data.iloc[0].get('cut', ''),
            'purchase_price': item_data.iloc[0].get('purchase_price', 0),
            'market_value': item_data.iloc[0].get('market_value', 0),
            'status': item_data.iloc[0].get('status', ''),
            'status_color': get_inventory_status_color(str(item_data.iloc[0].get('status', ''))),
            'location': item_data.iloc[0].get('location', ''),
            'purchase_date': item_data.iloc[0].get('purchase_date', ''),
            'notes': item_data.iloc[0].get('notes', ''),
            'rough_id': item_data.iloc[0].get('rough_id', ''),
            'kapan_no': item_data.iloc[0].get('kapan_no', '')
        }
        
        return render_template('inventory_item_details.html', item=item, rough_stone=rough_stone)
    except Exception as e:
        flash(f'Error loading item details: {str(e)}', 'error')
        return redirect(url_for('inventory'))

@app.route('/edit_inventory_item/<item_id>', methods=['GET', 'POST'])
def edit_inventory_item(item_id):
    if request.method == 'GET':
        try:
            # Load inventory data
            if not os.path.exists(INVENTORY_FILE):
                flash('Inventory file not found', 'error')
                return redirect(url_for('inventory'))
            
            inventory_df = pd.read_excel(INVENTORY_FILE)
            
            # Find the item
            item_data = inventory_df[inventory_df['id'] == item_id]
            
            if item_data.empty:
                flash('Item not found', 'error')
                return redirect(url_for('inventory'))
            
            # Prepare item for display
            item = {
                'id': item_data.iloc[0].get('id', ''),
                'description': item_data.iloc[0].get('description', ''),
                'shape': item_data.iloc[0].get('shape', ''),
                'carats': item_data.iloc[0].get('carats', 0),
                'color': item_data.iloc[0].get('color', ''),
                'clarity': item_data.iloc[0].get('clarity', ''),
                'cut': item_data.iloc[0].get('cut', ''),
                'purchase_price': item_data.iloc[0].get('purchase_price', 0),
                'market_value': item_data.iloc[0].get('market_value', 0),
                'status': item_data.iloc[0].get('status', ''),
                'location': item_data.iloc[0].get('location', ''),
                'notes': item_data.iloc[0].get('notes', ''),
                'rough_id': item_data.iloc[0].get('rough_id', ''),
                'kapan_no': item_data.iloc[0].get('kapan_no', '')
            }
            
            # Load rough inventory for dropdown
            rough_inventory_options = []
            if os.path.exists(ROUGH_INVENTORY_FILE):
                rough_inventory_df = pd.read_excel(ROUGH_INVENTORY_FILE)
                for _, row in rough_inventory_df.iterrows():
                    if row.get('status') != 'Processed' and row.get('status') != 'Sold' or row.get('id') == item['rough_id']:
                        rough_inventory_options.append({
                            'id': row.get('id', ''),
                            'description': row.get('description', ''),
                            'weight': row.get('weight', 0),
                            'pieces': row.get('pieces', 1)
                        })
            
            return render_template('edit_inventory_item.html', item=item, rough_inventory_options=rough_inventory_options)
        except Exception as e:
            flash(f'Error loading item for editing: {str(e)}', 'error')
            return redirect(url_for('inventory'))
    else:  # POST request
        try:
            # Get form data
            description = request.form.get('description')
            shape = request.form.get('shape')
            carats = float(request.form.get('carats'))
            color = request.form.get('color')
            clarity = request.form.get('clarity')
            cut = request.form.get('cut')
            purchase_price = float(request.form.get('purchase_price'))
            market_value = float(request.form.get('market_value'))
            status = request.form.get('status')
            location = request.form.get('location')
            notes = request.form.get('notes')
            rough_id = request.form.get('rough_id')
            kapan_no = request.form.get('kapan_no')
            
            # Load inventory data
            if not os.path.exists(INVENTORY_FILE):
                flash('Inventory file not found', 'error')
                return redirect(url_for('inventory'))
            
            inventory_df = pd.read_excel(INVENTORY_FILE)
            
            # Find the item
            item_index = inventory_df[inventory_df['id'] == item_id].index
            
            if len(item_index) == 0:
                flash('Item not found', 'error')
                return redirect(url_for('inventory'))
            
            # Get the previous rough_id to check if it changed
            previous_rough_id = inventory_df.loc[item_index[0], 'rough_id'] if 'rough_id' in inventory_df.columns else None
            
            # Update the item
            inventory_df.loc[item_index[0], 'description'] = description
            inventory_df.loc[item_index[0], 'shape'] = shape
            inventory_df.loc[item_index[0], 'carats'] = carats
            inventory_df.loc[item_index[0], 'color'] = color
            inventory_df.loc[item_index[0], 'clarity'] = clarity
            inventory_df.loc[item_index[0], 'cut'] = cut
            inventory_df.loc[item_index[0], 'purchase_price'] = purchase_price
            inventory_df.loc[item_index[0], 'market_value'] = market_value
            inventory_df.loc[item_index[0], 'status'] = status
            inventory_df.loc[item_index[0], 'location'] = location
            inventory_df.loc[item_index[0], 'notes'] = notes
            
            # Ensure rough_id column exists
            if 'rough_id' not in inventory_df.columns:
                inventory_df['rough_id'] = ''
            
            inventory_df.loc[item_index[0], 'rough_id'] = rough_id
            
            # Save to Excel
            inventory_df.to_excel(INVENTORY_FILE, index=False)
            
            # Update rough stone status if a rough stone was selected or changed
            if rough_id and rough_id != previous_rough_id:
                if os.path.exists(ROUGH_INVENTORY_FILE):
                    rough_inventory_df = pd.read_excel(ROUGH_INVENTORY_FILE)
                    rough_item_index = rough_inventory_df[rough_inventory_df['id'] == rough_id].index
                    
                    if len(rough_item_index) > 0:
                        # Check if all carats from rough stone have been processed
                        rough_weight = rough_inventory_df.loc[rough_item_index[0], 'weight']
                        
                        # Get all polished diamonds from this rough stone
                        polished_diamonds = inventory_df[inventory_df['rough_id'] == rough_id]
                        total_polished_carats = polished_diamonds['carats'].sum()
                        
                        # If more than 90% of the rough weight has been processed, mark as processed
                        if total_polished_carats >= (rough_weight * 0.9):
                            rough_inventory_df.loc[rough_item_index[0], 'status'] = 'Processed'
                            rough_inventory_df.to_excel(ROUGH_INVENTORY_FILE, index=False)
                            flash('Rough stone marked as processed as most of its weight has been converted to polished diamonds.', 'info')
            
            flash('Inventory item updated successfully!', 'success')
            return redirect(url_for('inventory_item_details', item_id=item_id))
        except Exception as e:
            flash(f'Error updating inventory item: {str(e)}', 'error')
            return redirect(url_for('inventory_item_details', item_id=item_id))

@app.route('/delete_inventory_item', methods=['POST'])
def delete_inventory_item():
    try:
        data = request.get_json()
        item_id = data.get('item_id')
        
        # Load inventory data
        if os.path.exists(INVENTORY_FILE):
            inventory_df = pd.read_excel(INVENTORY_FILE)
            
            # Find and remove the item
            inventory_df = inventory_df[inventory_df['id'] != item_id]
            
            # Save updated inventory
            inventory_df.to_excel(INVENTORY_FILE, index=False)
            
            return jsonify({'success': True, 'message': 'Item deleted successfully'})
        else:
            return jsonify({'success': False, 'message': 'Inventory file not found'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/upload_inventory', methods=['POST'])
def upload_inventory():
    try:
        if 'inventory_file' not in request.files:
            flash('No file part', 'error')
            return redirect(url_for('inventory'))
        
        file = request.files['inventory_file']
        
        if file.filename == '':
            flash('No selected file', 'error')
            return redirect(url_for('inventory'))
        
        if file and file.filename.endswith(('.xlsx', '.xls')):
            # Read the uploaded Excel file
            uploaded_df = pd.read_excel(file)
            
            # Check if the file has the required columns
            required_columns = ['description', 'shape', 'carats', 'color', 'clarity', 'cut', 
                               'purchase_price', 'market_value', 'status']
            
            missing_columns = [col for col in required_columns if col not in uploaded_df.columns]
            
            if missing_columns:
                flash(f'Missing required columns: {", ".join(missing_columns)}', 'error')
                return redirect(url_for('inventory'))
            
            # Load existing inventory
            if os.path.exists(INVENTORY_FILE):
                inventory_df = pd.read_excel(INVENTORY_FILE)
            else:
                inventory_df = pd.DataFrame(columns=['id', 'description', 'shape', 'carats', 'color', 'clarity', 'cut', 
                                                  'purchase_price', 'market_value', 'status', 'location', 'purchase_date', 'notes'])
            
            # Process each row in the uploaded file
            successful_imports = 0
            for _, row in uploaded_df.iterrows():
                try:
                    # Generate a unique ID for each new item
                    item_id = f"D{int(time.time())}{successful_imports}"
                    
                    # Create new item with required fields
                    new_item = {
                        'id': item_id,
                        'description': row.get('description', ''),
                        'shape': row.get('shape', ''),
                        'carats': float(row.get('carats', 0)),
                        'color': row.get('color', ''),
                        'clarity': row.get('clarity', ''),
                        'cut': row.get('cut', ''),
                        'purchase_price': float(row.get('purchase_price', 0)),
                        'market_value': float(row.get('market_value', 0)),
                        'status': row.get('status', 'In Stock'),
                        'location': row.get('location', ''),
                        'purchase_date': pd.Timestamp.now().strftime('%Y-%m-%d'),
                        'notes': row.get('notes', '')
                    }
                    
                    # Add to DataFrame
                    inventory_df = pd.concat([inventory_df, pd.DataFrame([new_item])], ignore_index=True)
                    successful_imports += 1
                    
                    # Add a small delay to ensure unique timestamps for IDs
                    time.sleep(0.01)
                    
                except Exception as e:
                    continue
            
            # Save to Excel
            inventory_df.to_excel(INVENTORY_FILE, index=False)
            
            flash(f'Successfully imported {successful_imports} inventory items!', 'success')
            return redirect(url_for('inventory'))
        else:
            flash('Invalid file format. Please upload an Excel file (.xlsx or .xls)', 'error')
            return redirect(url_for('inventory'))
    except Exception as e:
        flash(f'Error uploading inventory: {str(e)}', 'error')
        return redirect(url_for('inventory'))

@app.route('/download_inventory_template')
def download_inventory_template():
    try:
        # Create a template DataFrame with the required columns
        template_df = pd.DataFrame(columns=['description', 'shape', 'carats', 'color', 'clarity', 'cut', 
                                          'purchase_price', 'market_value', 'status', 'location', 'notes'])
        
        # Add a sample row to help users understand the format
        sample_row = {
            'description': 'Sample Diamond',
            'shape': 'Round',
            'carats': 1.25,
            'color': 'D',
            'clarity': 'VVS1',
            'cut': 'Excellent',
            'purchase_price': 100000,
            'market_value': 120000,
            'status': 'In Stock',
            'location': 'Safe Box 1',
            'notes': 'Sample notes'
        }
        template_df = pd.concat([template_df, pd.DataFrame([sample_row])], ignore_index=True)
        
        # Create a BytesIO object to store the Excel file
        output = io.BytesIO()
        
        # Write the DataFrame to the BytesIO object
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            template_df.to_excel(writer, index=False)
        
        # Seek to the beginning of the BytesIO object
        output.seek(0)
        
        # Return the Excel file as a response
        return send_file(
            output,
            as_attachment=True,
            download_name='inventory_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f'Error generating template: {str(e)}', 'error')
        return redirect(url_for('inventory'))

@app.route('/rough_inventory')
def rough_inventory():
    try:
        # Load rough inventory data
        if os.path.exists(ROUGH_INVENTORY_FILE):
            rough_inventory_df = pd.read_excel(ROUGH_INVENTORY_FILE)
        else:
            rough_inventory_df = pd.DataFrame(columns=['id', 'lot_id', 'description', 'source', 'origin', 
                                                    'weight', 'pieces', 'purchase_price', 'purchase_date', 
                                                    'status', 'location', 'notes', 'image_path', 'rough_id', 'kapan_no',
                                                    'shape_category'])
            rough_inventory_df.to_excel(ROUGH_INVENTORY_FILE, index=False)
        
        # Apply filters if provided
        status = request.args.get('status')
        source = request.args.get('source')
        min_weight = request.args.get('min_weight')
        max_weight = request.args.get('max_weight')
        min_price = request.args.get('min_price')
        max_price = request.args.get('max_price')
        is_bulk = request.args.get('is_bulk')
        
        filtered_df = rough_inventory_df.copy()
        
        if status:
            filtered_df = filtered_df[filtered_df['status'] == status]
        if source:
            filtered_df = filtered_df[filtered_df['source'] == source]
        if min_weight:
            filtered_df = filtered_df[filtered_df['weight'] >= float(min_weight)]
        if max_weight:
            filtered_df = filtered_df[filtered_df['weight'] <= float(max_weight)]
        if min_price:
            filtered_df = filtered_df[filtered_df['purchase_price'] >= float(min_price)]
        if max_price:
            filtered_df = filtered_df[filtered_df['purchase_price'] <= float(max_price)]
        if is_bulk:
            if is_bulk == 'yes':
                filtered_df = filtered_df[filtered_df['pieces'] > 1]
            elif is_bulk == 'no':
                filtered_df = filtered_df[filtered_df['pieces'] == 1]
        
        # Calculate totals
        total_items = len(filtered_df)
        total_value = filtered_df['purchase_price'].sum() if not filtered_df.empty else 0
        total_weight = filtered_df['weight'].sum() if not filtered_df.empty else 0
        total_pieces = filtered_df['pieces'].sum() if not filtered_df.empty else 0
        
        # Prepare rough inventory items for display
        rough_inventory = []
        for _, row in filtered_df.iterrows():
            item = {
                'id': row.get('id', ''),
                'lot_id': row.get('lot_id', ''),
                'description': row.get('description', ''),
                'source': row.get('source', ''),
                'origin': row.get('origin', ''),
                'weight': row.get('weight', 0),
                'pieces': row.get('pieces', 1),
                'purchase_price': row.get('purchase_price', 0),
                'purchase_date': row.get('purchase_date', ''),
                'status': row.get('status', ''),
                'status_color': get_inventory_status_color(str(row.get('status', ''))),
                'location': row.get('location', ''),
                'notes': row.get('notes', ''),
                'image_path': row.get('image_path', '')
            }
            rough_inventory.append(item)
        
        # Prepare chart data
        source_counts = filtered_df['source'].value_counts() if not filtered_df.empty else pd.Series()
        source_labels = source_counts.index.tolist()
        source_data = source_counts.values.tolist()
        
        status_counts = filtered_df['status'].value_counts() if not filtered_df.empty else pd.Series()
        status_labels = status_counts.index.tolist()
        status_data = status_counts.values.tolist()
        
        return render_template('rough_inventory.html', 
                             rough_inventory=rough_inventory,
                             total_items=total_items,
                             total_value=total_value,
                             total_weight=total_weight,
                             total_pieces=total_pieces,
                             source_labels=source_labels,
                             source_data=source_data,
                             status_labels=status_labels,
                             status_data=status_data)
    except Exception as e:
        flash(f'Error loading rough inventory: {str(e)}', 'error')
        return render_template('rough_inventory.html', 
                             rough_inventory=[],
                             total_items=0,
                             total_value=0,
                             total_weight=0,
                             total_pieces=0,
                             source_labels=[],
                             source_data=[],
                             status_labels=[],
                             status_data=[])

@app.route('/add_rough_inventory_item', methods=['POST'])
def add_rough_inventory_item():
    try:
        # Get form data
        description = request.form.get('description')
        source = request.form.get('source')
        origin = request.form.get('origin')
        weight = float(request.form.get('weight'))
        pieces = int(request.form.get('pieces', 1))
        purchase_price = float(request.form.get('purchase_price'))
        status = request.form.get('status')
        location = request.form.get('location')
        notes = request.form.get('notes')
        rough_id = request.form.get('rough_id')
        kapan_no = request.form.get('kapan_no')
        
        # Handle image upload if provided
        image_path = ''
        if 'item_image' in request.files:
            image = request.files['item_image']
            if image.filename != '':
                # Create images directory if it doesn't exist
                images_dir = os.path.join('static', 'images', 'rough')
                os.makedirs(images_dir, exist_ok=True)
                
                # Generate a unique filename
                filename = f"rough_{int(time.time())}_{secure_filename(image.filename)}"
                image_path = os.path.join(images_dir, filename)
                
                # Save the image
                image.save(image_path)
                
                # Store the relative path
                image_path = image_path.replace('\\', '/')
        
        # Load existing rough inventory
        if os.path.exists(ROUGH_INVENTORY_FILE):
            rough_inventory_df = pd.read_excel(ROUGH_INVENTORY_FILE)
        else:
            rough_inventory_df = pd.DataFrame(columns=['id', 'lot_id', 'description', 'source', 'origin', 
                                                    'weight', 'pieces', 'purchase_price', 'purchase_date', 
                                                    'status', 'location', 'notes', 'image_path', 'rough_id', 'kapan_no',
                                                    'shape_category'])
        
        # Generate a unique ID and lot ID
        item_id = f"R{int(time.time())}"
        lot_id = f"LOT-{int(time.time())[-4:]}" if pieces > 1 else ""
        
        # Create new item
        new_item = {
            'id': item_id,
            'lot_id': lot_id,
            'description': description,
            'source': source,
            'origin': origin,
            'weight': weight,
            'pieces': pieces,
            'purchase_price': purchase_price,
            'purchase_date': pd.Timestamp.now().strftime('%Y-%m-%d'),
            'status': status,
            'location': location,
            'notes': notes,
            'image_path': image_path,
            'rough_id': rough_id,
            'kapan_no': kapan_no,
            'shape_category': get_shape_from_form(request.form, 'rough')
        }
        
        # Add to DataFrame
        rough_inventory_df = pd.concat([rough_inventory_df, pd.DataFrame([new_item])], ignore_index=True)
        
        # Save to Excel
        rough_inventory_df.to_excel(ROUGH_INVENTORY_FILE, index=False)
        
        flash('Rough inventory item added successfully!', 'success')
        return redirect(url_for('rough_inventory'))
    except Exception as e:
        flash(f'Error adding rough inventory item: {str(e)}', 'error')
        return redirect(url_for('rough_inventory'))

@app.route('/rough_inventory_item_details/<item_id>')
def rough_inventory_item_details(item_id):
    try:
        # Load rough inventory data
        if not os.path.exists(ROUGH_INVENTORY_FILE):
            flash('Rough inventory file not found', 'error')
            return redirect(url_for('rough_inventory'))
        
        rough_inventory_df = pd.read_excel(ROUGH_INVENTORY_FILE)
        
        # Find the item
        item_data = rough_inventory_df[rough_inventory_df['id'] == item_id]
        
        if item_data.empty:
            flash('Item not found', 'error')
            return redirect(url_for('rough_inventory'))
        
        # Prepare item for display
        item = {
            'id': item_data.iloc[0].get('id', ''),
            'lot_id': item_data.iloc[0].get('lot_id', ''),
            'description': item_data.iloc[0].get('description', ''),
            'source': item_data.iloc[0].get('source', ''),
            'origin': item_data.iloc[0].get('origin', ''),
            'weight': item_data.iloc[0].get('weight', 0),
            'pieces': item_data.iloc[0].get('pieces', 1),
            'purchase_price': item_data.iloc[0].get('purchase_price', 0),
            'purchase_date': item_data.iloc[0].get('purchase_date', ''),
            'status': item_data.iloc[0].get('status', ''),
            'status_color': get_inventory_status_color(str(item_data.iloc[0].get('status', ''))),
            'location': item_data.iloc[0].get('location', ''),
            'notes': item_data.iloc[0].get('notes', ''),
            'image_path': item_data.iloc[0].get('image_path', '')
        }
        
        # Get related polished diamonds if any
        related_diamonds = []
        if os.path.exists(INVENTORY_FILE):
            inventory_df = pd.read_excel(INVENTORY_FILE)
            if 'rough_id' in inventory_df.columns:
                related_items = inventory_df[inventory_df['rough_id'] == item_id]
                
                for _, row in related_items.iterrows():
                    diamond = {
                        'id': row.get('id', ''),
                        'description': row.get('description', ''),
                        'shape': row.get('shape', ''),
                        'carats': row.get('carats', 0),
                        'status': row.get('status', ''),
                        'status_color': get_inventory_status_color(str(row.get('status', '')))
                    }
                    related_diamonds.append(diamond)
        
        return render_template('rough_inventory_item_details.html', item=item, related_diamonds=related_diamonds)
    except Exception as e:
        flash(f'Error loading item details: {str(e)}', 'error')
        return redirect(url_for('rough_inventory'))

@app.route('/edit_rough_inventory_item/<item_id>', methods=['GET', 'POST'])
def edit_rough_inventory_item(item_id):
    if request.method == 'GET':
        try:
            # Load rough inventory data
            if not os.path.exists(ROUGH_INVENTORY_FILE):
                flash('Rough inventory file not found', 'error')
                return redirect(url_for('rough_inventory'))
            
            rough_inventory_df = pd.read_excel(ROUGH_INVENTORY_FILE)
            
            # Find the item
            item_data = rough_inventory_df[rough_inventory_df['id'] == item_id]
            
            if item_data.empty:
                flash('Item not found', 'error')
                return redirect(url_for('rough_inventory'))
            
            # Prepare item for display
            item = {
                'id': item_data.iloc[0].get('id', ''),
                'lot_id': item_data.iloc[0].get('lot_id', ''),
                'description': item_data.iloc[0].get('description', ''),
                'source': item_data.iloc[0].get('source', ''),
                'origin': item_data.iloc[0].get('origin', ''),
                'weight': item_data.iloc[0].get('weight', 0),
                'pieces': item_data.iloc[0].get('pieces', 1),
                'purchase_price': item_data.iloc[0].get('purchase_price', 0),
                'purchase_date': item_data.iloc[0].get('purchase_date', ''),
                'status': item_data.iloc[0].get('status', ''),
                'location': item_data.iloc[0].get('location', ''),
                'notes': item_data.iloc[0].get('notes', ''),
                'image_path': item_data.iloc[0].get('image_path', '')
            }
            
            return render_template('edit_rough_inventory_item.html', item=item)
        except Exception as e:
            flash(f'Error loading item for editing: {str(e)}', 'error')
            return redirect(url_for('rough_inventory'))
    else:  # POST request
        try:
            # Get form data
            description = request.form.get('description')
            source = request.form.get('source')
            origin = request.form.get('origin')
            weight = float(request.form.get('weight'))
            pieces = int(request.form.get('pieces', 1))
            purchase_price = float(request.form.get('purchase_price'))
            status = request.form.get('status')
            location = request.form.get('location')
            notes = request.form.get('notes')
            rough_id = request.form.get('rough_id')
            kapan_no = request.form.get('kapan_no')
            
            # Load rough inventory data
            if not os.path.exists(ROUGH_INVENTORY_FILE):
                flash('Rough inventory file not found', 'error')
                return redirect(url_for('rough_inventory'))
            
            rough_inventory_df = pd.read_excel(ROUGH_INVENTORY_FILE)
            
            # Find the item
            item_index = rough_inventory_df[rough_inventory_df['id'] == item_id].index
            
            if len(item_index) == 0:
                flash('Item not found', 'error')
                return redirect(url_for('rough_inventory'))
            
            # Handle image upload if provided
            image_path = rough_inventory_df.loc[item_index[0], 'image_path']
            if 'item_image' in request.files:
                image = request.files['item_image']
                if image.filename != '':
                    # Create images directory if it doesn't exist
                    images_dir = os.path.join('static', 'images', 'rough')
                    os.makedirs(images_dir, exist_ok=True)
                    
                    # Generate a unique filename
                    filename = f"rough_{int(time.time())}_{secure_filename(image.filename)}"
                    new_image_path = os.path.join(images_dir, filename)
                    
                    # Save the image
                    image.save(new_image_path)
                    
                    # Store the relative path
                    image_path = new_image_path.replace('\\', '/')
            
            # Update the item
            rough_inventory_df.loc[item_index[0], 'description'] = description
            rough_inventory_df.loc[item_index[0], 'source'] = source
            rough_inventory_df.loc[item_index[0], 'origin'] = origin
            rough_inventory_df.loc[item_index[0], 'weight'] = weight
            rough_inventory_df.loc[item_index[0], 'pieces'] = pieces
            rough_inventory_df.loc[item_index[0], 'purchase_price'] = purchase_price
            rough_inventory_df.loc[item_index[0], 'status'] = status
            rough_inventory_df.loc[item_index[0], 'location'] = location
            rough_inventory_df.loc[item_index[0], 'notes'] = notes
            rough_inventory_df.loc[item_index[0], 'image_path'] = image_path
            
            # Update lot_id based on pieces
            if pieces > 1 and not rough_inventory_df.loc[item_index[0], 'lot_id']:
                rough_inventory_df.loc[item_index[0], 'lot_id'] = f"LOT-{str(int(time.time()))[-4:]}"
            elif pieces == 1:
                rough_inventory_df.loc[item_index[0], 'lot_id'] = ""
            
            rough_inventory_df.loc[item_index[0], 'rough_id'] = rough_id
            rough_inventory_df.loc[item_index[0], 'kapan_no'] = kapan_no
            rough_inventory_df.loc[item_index[0], 'shape_category'] = get_shape_from_form(request.form, 'rough')
            
            # Save to Excel
            rough_inventory_df.to_excel(ROUGH_INVENTORY_FILE, index=False)
            
            flash('Rough inventory item updated successfully!', 'success')
            return redirect(url_for('rough_inventory_item_details', item_id=item_id))
        except Exception as e:
            flash(f'Error updating rough inventory item: {str(e)}', 'error')
            return redirect(url_for('rough_inventory_item_details', item_id=item_id))

@app.route('/delete_rough_inventory_item', methods=['POST'])
def delete_rough_inventory_item():
    try:
        data = request.get_json()
        item_id = data.get('item_id')
        
        # Load rough inventory data
        if os.path.exists(ROUGH_INVENTORY_FILE):
            rough_inventory_df = pd.read_excel(ROUGH_INVENTORY_FILE)
            
            # Find the item
            item_data = rough_inventory_df[rough_inventory_df['id'] == item_id]
            
            if not item_data.empty:
                # Check if there are related polished diamonds
                has_related_diamonds = False
                if os.path.exists(INVENTORY_FILE):
                    inventory_df = pd.read_excel(INVENTORY_FILE)
                    if 'rough_id' in inventory_df.columns:
                        related_items = inventory_df[inventory_df['rough_id'] == item_id]
                        has_related_diamonds = not related_items.empty
                
                if has_related_diamonds:
                    return jsonify({'success': False, 'message': 'Cannot delete rough stone with related polished diamonds'})
                
                # Delete the image if it exists
                image_path = item_data.iloc[0].get('image_path', '')
                if image_path and os.path.exists(image_path):
                    os.remove(image_path)
                
                # Remove the item
                rough_inventory_df = rough_inventory_df[rough_inventory_df['id'] != item_id]
                
                # Save updated inventory
                rough_inventory_df.to_excel(ROUGH_INVENTORY_FILE, index=False)
                
                return jsonify({'success': True, 'message': 'Item deleted successfully'})
            else:
                return jsonify({'success': False, 'message': 'Item not found'})
        else:
            return jsonify({'success': False, 'message': 'Rough inventory file not found'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/download_rough_inventory_template')
def download_rough_inventory_template():
    try:
        # Create a template DataFrame with the required columns
        template_df = pd.DataFrame(columns=['description', 'source', 'origin', 'weight', 'pieces', 
                                          'purchase_price', 'status', 'location', 'notes'])
        
        # Add a sample row to help users understand the format
        sample_row = {
            'description': 'Sample Rough Diamond',
            'source': 'Mine Direct',
            'origin': 'South Africa',
            'weight': 5.75,
            'pieces': 3,
            'purchase_price': 250000,
            'status': 'In Stock',
            'location': 'Vault 2',
            'notes': 'Good quality rough with minimal inclusions'
        }
        template_df = pd.concat([template_df, pd.DataFrame([sample_row])], ignore_index=True)
        
        # Create a BytesIO object to store the Excel file
        output = io.BytesIO()
        
        # Write the DataFrame to the BytesIO object
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            template_df.to_excel(writer, index=False)
        
        # Seek to the beginning of the BytesIO object
        output.seek(0)
        
        # Return the Excel file as a response
        return send_file(
            output,
            as_attachment=True,
            download_name='rough_inventory_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f'Error generating template: {str(e)}', 'error')
        return redirect(url_for('rough_inventory'))

@app.route('/upload_rough_inventory', methods=['POST'])
def upload_rough_inventory():
    try:
        if 'inventory_file' not in request.files:
            flash('No file part', 'error')
            return redirect(url_for('rough_inventory'))
        
        file = request.files['inventory_file']
        
        if file.filename == '':
            flash('No selected file', 'error')
            return redirect(url_for('rough_inventory'))
        
        if file and file.filename.endswith(('.xlsx', '.xls')):
            # Read the uploaded Excel file
            uploaded_df = pd.read_excel(file)
            
            # Check if the file has the required columns
            required_columns = ['description', 'source', 'weight', 'pieces', 'purchase_price', 'status']
            
            missing_columns = [col for col in required_columns if col not in uploaded_df.columns]
            
            if missing_columns:
                flash(f'Missing required columns: {", ".join(missing_columns)}', 'error')
                return redirect(url_for('rough_inventory'))
            
            # Load existing rough inventory
            if os.path.exists(ROUGH_INVENTORY_FILE):
                rough_inventory_df = pd.read_excel(ROUGH_INVENTORY_FILE)
            else:
                rough_inventory_df = pd.DataFrame(columns=['id', 'lot_id', 'description', 'source', 'origin', 
                                                        'weight', 'pieces', 'purchase_price', 'purchase_date', 
                                                        'status', 'location', 'notes', 'image_path', 'rough_id', 'kapan_no',
                                                        'shape_category'])
            
            # Process each row in the uploaded file
            successful_imports = 0
            for _, row in uploaded_df.iterrows():
                try:
                    # Generate a unique ID
                    item_id = f"R{int(time.time())}{successful_imports}"
                    
                    # Determine if it's a bulk lot
                    pieces = int(row.get('pieces', 1))
                    lot_id = f"LOT-{str(int(time.time()))[-4:]}" if pieces > 1 else ""
                    
                    # Create new item with required fields
                    new_item = {
                        'id': item_id,
                        'lot_id': lot_id,
                        'description': row.get('description', ''),
                        'source': row.get('source', ''),
                        'origin': row.get('origin', ''),
                        'weight': float(row.get('weight', 0)),
                        'pieces': pieces,
                        'purchase_price': float(row.get('purchase_price', 0)),
                        'purchase_date': pd.Timestamp.now().strftime('%Y-%m-%d'),
                        'status': row.get('status', 'In Stock'),
                        'location': row.get('location', ''),
                        'notes': row.get('notes', ''),
                        'image_path': ''
                    }
                    
                    # Add to DataFrame
                    rough_inventory_df = pd.concat([rough_inventory_df, pd.DataFrame([new_item])], ignore_index=True)
                    successful_imports += 1
                    
                    # Add a small delay to ensure unique timestamps for IDs
                    time.sleep(0.01)
                    
                except Exception as e:
                    continue
            
            # Save to Excel
            rough_inventory_df.to_excel(ROUGH_INVENTORY_FILE, index=False)
            
            flash(f'Successfully imported {successful_imports} rough inventory items!', 'success')
            return redirect(url_for('rough_inventory'))
        else:
            flash('Invalid file format. Please upload an Excel file (.xlsx or .xls)', 'error')
            return redirect(url_for('rough_inventory'))
    except Exception as e:
        flash(f'Error uploading rough inventory: {str(e)}', 'error')
        return redirect(url_for('rough_inventory'))

@app.route('/dashboard')
def dashboard():
    """
    Display a dashboard with key metrics and visualizations.
    """
    try:
        # Prepare data for dashboard
        inventory_stats = {}
        sales_stats = {}
        purchase_stats = {}
        
        # Get inventory statistics if file exists
        if os.path.exists(INVENTORY_FILE):
            inventory_df = pd.read_excel(INVENTORY_FILE)
            inventory_stats = {
                'total_items': len(inventory_df),
                'total_carats': inventory_df['carats'].sum() if 'carats' in inventory_df.columns else 0,
                'total_value': inventory_df['price'].sum() if 'price' in inventory_df.columns else 0
            }
        
        # Get rough inventory statistics if file exists
        rough_inventory_stats = {}
        if os.path.exists(ROUGH_INVENTORY_FILE):
            rough_df = pd.read_excel(ROUGH_INVENTORY_FILE)
            rough_inventory_stats = {
                'total_items': len(rough_df),
                'total_weight': rough_df['weight'].sum() if 'weight' in rough_df.columns else 0,
                'total_value': rough_df['purchase_price'].sum() if 'purchase_price' in rough_df.columns else 0
            }
        
        # Get sales statistics if file exists
        if os.path.exists(SALES_FILE):
            sales_df = pd.read_excel(SALES_FILE)
            sales_stats = {
                'total_sales': len(sales_df),
                'total_amount': sales_df['Total Amount USD'].sum() if 'Total Amount USD' in sales_df.columns else 0
            }
        
        # Get purchase statistics if file exists
        if os.path.exists(PURCHASES_FILE):
            purchases_df = pd.read_excel(PURCHASES_FILE)
            purchase_stats = {
                'total_purchases': len(purchases_df),
                'total_amount': purchases_df['Total Amount USD'].sum() if 'Total Amount USD' in purchases_df.columns else 0
            }
        
        return render_template('dashboard.html', 
                              inventory_stats=inventory_stats,
                              rough_inventory_stats=rough_inventory_stats,
                              sales_stats=sales_stats,
                              purchase_stats=purchase_stats)
    except Exception as e:
        flash(f'Error loading dashboard: {str(e)}', 'error')
        return render_template('dashboard.html', 
                              inventory_stats={},
                              rough_inventory_stats={},
                              sales_stats={},
                              purchase_stats={})

@app.route('/sell', methods=['GET', 'POST'])
def sell():
    """
    Handle diamond sales transactions.
    """
    try:
        if request.method == 'POST':
            # Get the diamond type from the form
            diamond_type = request.form.get('diamond_type', 'polished')
            
            if diamond_type == 'polished':
                # Process polished diamond sale
                inventory_item_id = request.form.get('inventory_item')
                customer_name = request.form.get('customer_name')
                sale_price = float(request.form.get('sale_price'))
                sale_date = request.form.get('sale_date')
                notes = request.form.get('notes', '')
                
                # Update inventory status to 'Sold'
                if os.path.exists(INVENTORY_FILE):
                    inventory_df = pd.read_excel(INVENTORY_FILE)
                    item_index = inventory_df[inventory_df['id'] == inventory_item_id].index
                    
                    if len(item_index) > 0:
                        inventory_df.loc[item_index[0], 'status'] = 'Sold'
                        inventory_df.to_excel(INVENTORY_FILE, index=False)
                        
                        # Record the sale in sales.xlsx
                        if os.path.exists(os.path.join(DATA_DIR, 'sales.xlsx')):
                            sales_df = pd.read_excel(os.path.join(DATA_DIR, 'sales.xlsx'))
                        else:
                            sales_df = pd.DataFrame(columns=['id', 'item_id', 'customer_name', 'sale_price', 'sale_date', 'notes', 'diamond_type'])
                        
                        # Create new sale record
                        sale_id = f"S{int(time.time())}"
                        new_sale = {
                            'id': sale_id,
                            'item_id': inventory_item_id,
                            'customer_name': customer_name,
                            'sale_price': sale_price,
                            'sale_date': sale_date,
                            'notes': notes,
                            'diamond_type': 'polished'
                        }
                        
                        sales_df = pd.concat([sales_df, pd.DataFrame([new_sale])], ignore_index=True)
                        sales_df.to_excel(os.path.join(DATA_DIR, 'sales.xlsx'), index=False)
                        
                        flash('Polished diamond sale recorded successfully!', 'success')
                    else:
                        flash('Inventory item not found', 'error')
                else:
                    flash('Inventory file not found', 'error')
            
            elif diamond_type == 'rough':
                # Process rough diamond sale
                rough_item_id = request.form.get('rough_inventory_item')
                customer_name = request.form.get('customer_name')
                sale_price = float(request.form.get('sale_price'))
                sale_date = request.form.get('sale_date')
                notes = request.form.get('notes', '')
                
                # Update rough inventory status to 'Sold'
                if os.path.exists(ROUGH_INVENTORY_FILE):
                    rough_inventory_df = pd.read_excel(ROUGH_INVENTORY_FILE)
                    item_index = rough_inventory_df[rough_inventory_df['id'] == rough_item_id].index
                    
                    if len(item_index) > 0:
                        rough_inventory_df.loc[item_index[0], 'status'] = 'Sold'
                        rough_inventory_df.to_excel(ROUGH_INVENTORY_FILE, index=False)
                        
                        # Record the sale in sales.xlsx
                        if os.path.exists(os.path.join(DATA_DIR, 'sales.xlsx')):
                            sales_df = pd.read_excel(os.path.join(DATA_DIR, 'sales.xlsx'))
                        else:
                            sales_df = pd.DataFrame(columns=['id', 'item_id', 'customer_name', 'sale_price', 'sale_date', 'notes', 'diamond_type'])
                        
                        # Create new sale record
                        sale_id = f"S{int(time.time())}"
                        new_sale = {
                            'id': sale_id,
                            'item_id': rough_item_id,
                            'customer_name': customer_name,
                            'sale_price': sale_price,
                            'sale_date': sale_date,
                            'notes': notes,
                            'diamond_type': 'rough'
                        }
                        
                        sales_df = pd.concat([sales_df, pd.DataFrame([new_sale])], ignore_index=True)
                        sales_df.to_excel(os.path.join(DATA_DIR, 'sales.xlsx'), index=False)
                        
                        flash('Rough diamond sale recorded successfully!', 'success')
                    else:
                        flash('Rough inventory item not found', 'error')
                else:
                    flash('Rough inventory file not found', 'error')
            
            return redirect(url_for('sell'))
        
        # For GET requests, display the sell form
        # Load polished inventory for selection
        inventory_items = []
        if os.path.exists(INVENTORY_FILE):
            inventory_df = pd.read_excel(INVENTORY_FILE)
            inventory_df = inventory_df[inventory_df['status'] == 'In Stock']
            
            for _, row in inventory_df.iterrows():
                item = {
                    'id': row.get('id', ''),
                    'description': row.get('description', ''),
                    'shape': row.get('shape', ''),
                    'carats': row.get('carats', 0),
                    'color': row.get('color', ''),
                    'clarity': row.get('clarity', ''),
                    'market_value': row.get('market_value', 0)
                }
                inventory_items.append(item)
        
        # Load rough inventory for selection
        rough_inventory_items = []
        if os.path.exists(ROUGH_INVENTORY_FILE):
            rough_inventory_df = pd.read_excel(ROUGH_INVENTORY_FILE)
            rough_inventory_df = rough_inventory_df[rough_inventory_df['status'] == 'In Stock']
            
            for _, row in rough_inventory_df.iterrows():
                item = {
                    'id': row.get('id', ''),
                    'description': row.get('description', ''),
                    'weight': row.get('weight', 0),
                    'rough_id': row.get('rough_id', ''),
                    'kapan_no': row.get('kapan_no', ''),
                    'purchase_price': row.get('purchase_price', 0)
                }
                rough_inventory_items.append(item)
        
        return render_template('sell.html', 
                              inventory_items=inventory_items,
                              rough_inventory_items=rough_inventory_items,
                              today_date=datetime.now().strftime('%Y-%m-%d'))
    except Exception as e:
        flash(f'Error loading sell page: {str(e)}', 'error')
        return render_template('sell.html', inventory_items=[], rough_inventory_items=[])

@app.route('/records')
def records():
    """
    Display transaction records.
    """
    try:
        # Prepare data for records page
        purchases = []
        sales = []
        
        # Get purchase records if file exists
        if os.path.exists(PURCHASES_FILE):
            purchases_df = pd.read_excel(PURCHASES_FILE)
            for _, row in purchases_df.iterrows():
                purchase = {
                    'id': row.get('id', ''),
                    'date': row.get('Date', ''),
                    'party': row.get('Party', ''),
                    'description': row.get('Description', ''),
                    'carat': row.get('Carat', 0),
                    'amount_usd': row.get('Total Amount USD', 0),
                    'amount_inr': row.get('Total Amount INR', 0),
                    'payment_status': row.get('Payment Status', '')
                }
                purchases.append(purchase)
        
        # Get sales records if file exists
        if os.path.exists(SALES_FILE):
            sales_df = pd.read_excel(SALES_FILE)
            for _, row in sales_df.iterrows():
                sale = {
                    'id': row.get('id', ''),
                    'date': row.get('Date', ''),
                    'party': row.get('Party', ''),
                    'description': row.get('Description', ''),
                    'carat': row.get('Carat', 0),
                    'amount_usd': row.get('Total Amount USD', 0),
                    'amount_inr': row.get('Total Amount INR', 0),
                    'payment_status': row.get('Payment Status', '')
                }
                sales.append(sale)
        
        return render_template('records.html', purchases=purchases, sales=sales)
    except Exception as e:
        flash(f'Error loading records: {str(e)}', 'error')
        return render_template('records.html', purchases=[], sales=[])

@app.route('/payments')
def payments():
    """
    Display payment records.
    """
    try:
        # Prepare data for payments page
        payments_list = []
        
        # Get payment records if file exists
        if os.path.exists(PAYMENTS_FILE):
            payments_df = pd.read_excel(PAYMENTS_FILE)
            for _, row in payments_df.iterrows():
                payment = {
                    'id': row.get('id', ''),
                    'date': row.get('Date', ''),
                    'party': row.get('Party', ''),
                    'description': row.get('Description', ''),
                    'amount_usd': row.get('Amount USD', 0),
                    'amount_inr': row.get('Amount INR', 0),
                    'payment_method': row.get('Payment Method', ''),
                    'reference': row.get('Reference', '')
                }
                payments_list.append(payment)
        
        return render_template('payments.html', payments=payments_list)
    except Exception as e:
        flash(f'Error loading payments: {str(e)}', 'error')
        return render_template('payments.html', payments=[])

@app.route('/reports')
def reports():
    """
    Display business reports.
    """
    try:
        # Prepare data for reports page
        report_data = {
            'total_inventory_value': 0,
            'total_inventory_carats': 0,
            'total_rough_inventory_value': 0,
            'total_rough_inventory_weight': 0,
            'total_purchases': 0,
            'total_sales': 0,
            'profit_margin': 0
        }
        
        # Get inventory statistics if file exists
        if os.path.exists(INVENTORY_FILE):
            inventory_df = pd.read_excel(INVENTORY_FILE)
            report_data['total_inventory_value'] = inventory_df['price'].sum() if 'price' in inventory_df.columns else 0
            report_data['total_inventory_carats'] = inventory_df['carats'].sum() if 'carats' in inventory_df.columns else 0
        
        # Get rough inventory statistics if file exists
        if os.path.exists(ROUGH_INVENTORY_FILE):
            rough_df = pd.read_excel(ROUGH_INVENTORY_FILE)
            report_data['total_rough_inventory_value'] = rough_df['purchase_price'].sum() if 'purchase_price' in rough_df.columns else 0
            report_data['total_rough_inventory_weight'] = rough_df['weight'].sum() if 'weight' in rough_df.columns else 0
        
        # Get purchase statistics if file exists
        if os.path.exists(PURCHASES_FILE):
            purchases_df = pd.read_excel(PURCHASES_FILE)
            report_data['total_purchases'] = purchases_df['Total Amount USD'].sum() if 'Total Amount USD' in purchases_df.columns else 0
        
        # Get sales statistics if file exists
        if os.path.exists(SALES_FILE):
            sales_df = pd.read_excel(SALES_FILE)
            report_data['total_sales'] = sales_df['Total Amount USD'].sum() if 'Total Amount USD' in sales_df.columns else 0
        
        # Calculate profit margin
        if report_data['total_sales'] > 0 and report_data['total_purchases'] > 0:
            report_data['profit_margin'] = ((report_data['total_sales'] - report_data['total_purchases']) / report_data['total_purchases']) * 100
        
        return render_template('reports.html', report_data=report_data)
    except Exception as e:
        flash(f'Error loading reports: {str(e)}', 'error')
        return render_template('reports.html', report_data={})

@app.route('/add_rough_inventory')
def add_rough_inventory():
    return render_template('add_rough_inventory_item.html')

@app.route('/backup', methods=['GET', 'POST'])
def backup():
    """
    Handle backup and restore operations.
    """
    try:
        logger.info(f"Backup route accessed: {request.method}")
        
        if request.method == 'POST':
            action = request.form.get('action')
            logger.info(f"Backup action: {action}")
            
            # Check if this is an AJAX request
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            
            if action == 'create_backup':
                # Create a new backup
                backup_file = create_backup()
                if backup_file:
                    if is_ajax:
                        return jsonify({
                            'success': True,
                            'message': f'Backup created successfully: {os.path.basename(backup_file)}'
                        })
                    else:
                        flash(f'Backup created successfully: {os.path.basename(backup_file)}', 'success')
                else:
                    if is_ajax:
                        return jsonify({
                            'success': False,
                            'message': 'Error creating backup'
                        })
                    else:
                        flash('Error creating backup', 'error')
            
            elif action == 'restore_backup':
                # Restore from a selected backup
                backup_file = request.form.get('backup_file')
                if backup_file:
                    backup_path = os.path.join(BACKUP_DIR, backup_file)
                    if os.path.exists(backup_path):
                        if restore_from_backup(backup_path):
                            if is_ajax:
                                return jsonify({
                                    'success': True,
                                    'message': f'Data restored successfully from {backup_file}'
                                })
                            else:
                                flash(f'Data restored successfully from {backup_file}', 'success')
                        else:
                            if is_ajax:
                                return jsonify({
                                    'success': False,
                                    'message': f'Error restoring from {backup_file}'
                                })
                            else:
                                flash(f'Error restoring from {backup_file}', 'error')
                    else:
                        if is_ajax:
                            return jsonify({
                                'success': False,
                                'message': 'Backup file not found'
                            })
                        else:
                            flash('Backup file not found', 'error')
                else:
                    if is_ajax:
                        return jsonify({
                            'success': False,
                            'message': 'No backup file selected',
                            'errors': {'backup_file': 'Please select a backup file'}
                        })
                    else:
                        flash('No backup file selected', 'error')
        
        # Get list of available backups
        backup_files = []
        if os.path.exists(BACKUP_DIR):
            backup_files = sorted([f for f in os.listdir(BACKUP_DIR) 
                                if f.startswith('diamond_data_backup_') and f.endswith('.zip')],
                                key=lambda x: os.path.getmtime(os.path.join(BACKUP_DIR, x)),
                                reverse=True)
        
        return render_template('backup.html', backup_files=backup_files, data_dir=DATA_DIR)
    except Exception as e:
        # Log the error for debugging
        print(f"Error in backup route: {str(e)}")
        
        # Check if this is an AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': False,
                'message': f'An unexpected error occurred: {str(e)}'
            })
        
        # Show a user-friendly error message
        flash(f'An unexpected error occurred: {str(e)}', 'error')
        return render_template('backup.html', backup_files=[], data_dir=DATA_DIR)

# Function to validate data consistency across files
def validate_data_consistency():
    """
    Validate data consistency across files.
    Ensures that references between files are valid.
    Returns a list of inconsistencies found.
    """
    inconsistencies = []
    
    try:
        # Load all data files
        inventory_df = pd.DataFrame()
        if os.path.exists(INVENTORY_FILE):
            inventory_df = pd.read_excel(INVENTORY_FILE)
        
        rough_inventory_df = pd.DataFrame()
        if os.path.exists(ROUGH_INVENTORY_FILE):
            rough_inventory_df = pd.read_excel(ROUGH_INVENTORY_FILE)
        
        purchases_df = pd.DataFrame()
        if os.path.exists(PURCHASES_FILE):
            purchases_df = pd.read_excel(PURCHASES_FILE)
        
        sales_df = pd.DataFrame()
        if os.path.exists(SALES_FILE):
            sales_df = pd.read_excel(SALES_FILE)
        
        payments_df = pd.DataFrame()
        if os.path.exists(PAYMENTS_FILE):
            payments_df = pd.read_excel(PAYMENTS_FILE)
        
        # Check if item IDs in sales exist in inventory
        if not sales_df.empty and not inventory_df.empty and 'item_id' in sales_df.columns:
            for _, row in sales_df.iterrows():
                item_id = row.get('item_id')
                if item_id and item_id not in inventory_df['id'].values and item_id not in rough_inventory_df['id'].values:
                    inconsistencies.append(f"Item ID {item_id} in sales does not exist in inventory")
        
        # Check if rough IDs in inventory exist in rough inventory
        if not inventory_df.empty and not rough_inventory_df.empty and 'rough_id' in inventory_df.columns:
            for _, row in inventory_df.iterrows():
                rough_id = row.get('rough_id')
                if rough_id and rough_id not in rough_inventory_df['id'].values and rough_id != '':
                    inconsistencies.append(f"Rough ID {rough_id} in inventory does not exist in rough inventory")
        
        # Check if item IDs in purchases exist in inventory
        if not purchases_df.empty and not inventory_df.empty and 'item_id' in purchases_df.columns:
            for _, row in purchases_df.iterrows():
                item_id = row.get('item_id')
                diamond_type = row.get('diamond_type')
                
                if item_id and diamond_type == 'polished' and item_id not in inventory_df['id'].values:
                    inconsistencies.append(f"Item ID {item_id} in purchases (polished) does not exist in inventory")
                elif item_id and diamond_type == 'rough' and item_id not in rough_inventory_df['id'].values:
                    inconsistencies.append(f"Item ID {item_id} in purchases (rough) does not exist in rough inventory")
        
        return inconsistencies
    except Exception as e:
        inconsistencies.append(f"Error validating data consistency: {str(e)}")
        return inconsistencies

# Function to fix data inconsistencies
def fix_data_inconsistencies():
    """
    Fix data inconsistencies across files.
    Removes invalid references between files.
    Returns the number of inconsistencies fixed.
    """
    fixed_count = 0
    
    try:
        # Load all data files
        inventory_df = pd.DataFrame()
        if os.path.exists(INVENTORY_FILE):
            inventory_df = pd.read_excel(INVENTORY_FILE)
        
        rough_inventory_df = pd.DataFrame()
        if os.path.exists(ROUGH_INVENTORY_FILE):
            rough_inventory_df = pd.read_excel(ROUGH_INVENTORY_FILE)
        
        purchases_df = pd.DataFrame()
        if os.path.exists(PURCHASES_FILE):
            purchases_df = pd.read_excel(PURCHASES_FILE)
        
        sales_df = pd.DataFrame()
        if os.path.exists(SALES_FILE):
            sales_df = pd.read_excel(SALES_FILE)
        
        payments_df = pd.DataFrame()
        if os.path.exists(PAYMENTS_FILE):
            payments_df = pd.read_excel(PAYMENTS_FILE)
        
        # Fix invalid rough IDs in inventory
        if not inventory_df.empty and not rough_inventory_df.empty and 'rough_id' in inventory_df.columns:
            valid_rough_ids = set(rough_inventory_df['id'].values) if 'id' in rough_inventory_df.columns else set()
            for idx, row in inventory_df.iterrows():
                rough_id = row.get('rough_id')
                if rough_id and rough_id not in valid_rough_ids and rough_id != '':
                    inventory_df.at[idx, 'rough_id'] = ''
                    fixed_count += 1
            
            if fixed_count > 0:
                inventory_df.to_excel(INVENTORY_FILE, index=False)
                print(f"Fixed {fixed_count} invalid rough IDs in inventory")
        
        # Fix invalid item IDs in sales
        if not sales_df.empty and ('item_id' in sales_df.columns):
            valid_inventory_ids = set(inventory_df['id'].values) if not inventory_df.empty and 'id' in inventory_df.columns else set()
            valid_rough_ids = set(rough_inventory_df['id'].values) if not rough_inventory_df.empty and 'id' in rough_inventory_df.columns else set()
            valid_ids = valid_inventory_ids.union(valid_rough_ids)
            
            invalid_rows = []
            for idx, row in sales_df.iterrows():
                item_id = row.get('item_id')
                if item_id and item_id not in valid_ids:
                    invalid_rows.append(idx)
            
            if invalid_rows:
                sales_df = sales_df.drop(invalid_rows)
                sales_df.to_excel(SALES_FILE, index=False)
                fixed_count += len(invalid_rows)
                print(f"Removed {len(invalid_rows)} sales records with invalid item IDs")
        
        # Fix invalid item IDs in purchases
        if not purchases_df.empty and ('item_id' in purchases_df.columns) and ('diamond_type' in purchases_df.columns):
            valid_inventory_ids = set(inventory_df['id'].values) if not inventory_df.empty and 'id' in inventory_df.columns else set()
            valid_rough_ids = set(rough_inventory_df['id'].values) if not rough_inventory_df.empty and 'id' in rough_inventory_df.columns else set()
            
            invalid_rows = []
            for idx, row in purchases_df.iterrows():
                item_id = row.get('item_id')
                diamond_type = row.get('diamond_type')
                
                if item_id:
                    if diamond_type == 'polished' and item_id not in valid_inventory_ids:
                        invalid_rows.append(idx)
                    elif diamond_type == 'rough' and item_id not in valid_rough_ids:
                        invalid_rows.append(idx)
            
            if invalid_rows:
                purchases_df = purchases_df.drop(invalid_rows)
                purchases_df.to_excel(PURCHASES_FILE, index=False)
                fixed_count += len(invalid_rows)
                print(f"Removed {len(invalid_rows)} purchase records with invalid item IDs")
        
        return fixed_count
    except Exception as e:
        print(f"Error fixing data inconsistencies: {str(e)}")
        return fixed_count

# Function to schedule backups at regular intervals
def schedule_backups(interval_hours=24):
    """
    Schedule backups at regular intervals.
    Args:
        interval_hours: Interval in hours between backups.
    """
    def backup_task():
        while True:
            try:
                # Create a backup
                backup_file = create_backup()
                if backup_file:
                    print(f"Scheduled backup created: {os.path.basename(backup_file)}")
                else:
                    print("Failed to create scheduled backup")
                
                # Sleep for the specified interval
                time.sleep(interval_hours * 3600)
            except Exception as e:
                print(f"Error in scheduled backup: {str(e)}")
                # Sleep for a shorter interval before retrying
                time.sleep(3600)
    
    # Start the backup task in a background thread
    backup_thread = threading.Thread(target=backup_task, daemon=True)
    backup_thread.start()
    print(f"Scheduled backups every {interval_hours} hours")

@app.route('/debug/responsive', methods=['GET'])
def debug_responsive():
    """
    Debug route for responsive design testing.
    Only available in debug mode.
    """
    if not app.debug:
        return render_template('error.html', 
                              error_code=403, 
                              error_message="This debugging tool is only available in debug mode."), 403
    
    logger.info("Responsive design testing tool accessed")
    return render_template('debug_responsive.html')

if __name__ == '__main__':
    # Validate data consistency
    try:
        inconsistencies = validate_data_consistency()
        if inconsistencies:
            print("Data inconsistencies found:")
            for inconsistency in inconsistencies:
                print(f"  - {inconsistency}")
            
            # Fix data inconsistencies
            fixed_count = fix_data_inconsistencies()
            if fixed_count > 0:
                print(f"Fixed {fixed_count} data inconsistencies")
                
                # Validate again after fixing
                inconsistencies = validate_data_consistency()
                if inconsistencies:
                    print("Remaining data inconsistencies:")
                    for inconsistency in inconsistencies:
                        print(f"  - {inconsistency}")
                else:
                    print("All data inconsistencies fixed successfully")
        else:
            print("No data inconsistencies found")
    except Exception as e:
        print(f"Error validating data consistency: {str(e)}")
    
    # Fix data type inconsistencies
    try:
        if fix_data_types():
            print("Data type inconsistencies fixed successfully")
        else:
            print("Failed to fix data type inconsistencies")
    except Exception as e:
        print(f"Error fixing data type inconsistencies: {str(e)}")
    
    # Create a backup when the application starts
    try:
        backup_file = create_backup()
        if backup_file:
            print(f"Automatic backup created: {os.path.basename(backup_file)}")
        else:
            print("Failed to create automatic backup")
    except Exception as e:
        print(f"Error creating automatic backup: {str(e)}")
    
    # Schedule backups every 6 hours
    schedule_backups(interval_hours=6)
    
    # Run the application
    app.run(debug=True) 