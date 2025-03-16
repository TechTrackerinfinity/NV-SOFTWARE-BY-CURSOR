import pandas as pd
import os
from flask import current_app
import logging
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook

logger = logging.getLogger('diamond_app')

__all__ = ['fix_data_types', 'validate_data_consistency', 'enhance_excel_formatting']

def fix_data_types():
    """
    Fix data type inconsistencies in Excel files.
    Ensures that numeric columns are stored as appropriate numeric types.
    """
    try:
        data_dir = current_app.config['DATA_DIR']
        files_to_fix = {
            'inventory.xlsx': {
                'numeric_columns': ['carats', 'purchase_price', 'market_value']
            },
            'rough_inventory.xlsx': {
                'numeric_columns': ['weight', 'pieces', 'purchase_price'],
                'required_columns': ['rough_id', 'kapan_no', 'shape_category']
            },
            'purchases.xlsx': {
                'numeric_columns': ['amount', 'Carat', 'Price Per Carat', 'Total Amount']
            },
            'sales.xlsx': {
                'numeric_columns': ['Carat', 'Price Per Carat', 'Total Amount', 
                                  'carat', 'price_per_carat', 'total_amount_usd']
            },
            'payments.xlsx': {
                'numeric_columns': ['total_amount', 'paid_amount', 'pending_amount']
            }
        }
        
        for file_name, config in files_to_fix.items():
            file_path = os.path.join(data_dir, file_name)
            if os.path.exists(file_path):
                df = pd.read_excel(file_path)
                if not df.empty:
                    # Convert numeric columns to appropriate types
                    for col in config['numeric_columns']:
                        if col in df.columns:
                            df[col] = pd.to_numeric(df[col], errors='coerce')
                    
                    # Add required columns if specified
                    if 'required_columns' in config:
                        for col in config['required_columns']:
                            if col not in df.columns:
                                df[col] = ''
                    
                    # Save the fixed DataFrame
                    df.to_excel(file_path, index=False)
                    logger.info(f"Fixed data types in {file_name}")
        
        return True
    except Exception as e:
        logger.error(f"Error fixing data types: {str(e)}")
        return False

def validate_data_consistency():
    """
    Validate data consistency across Excel files.
    Checks for missing required columns, data type consistency, and referential integrity.
    """
    try:
        data_dir = current_app.config['DATA_DIR']
        validation_rules = {
            'inventory.xlsx': {
                'required_columns': ['item_id', 'carats', 'clarity', 'color', 'cut', 'price'],
                'numeric_columns': ['carats', 'price'],
                'non_empty_columns': ['item_id']
            },
            'rough_inventory.xlsx': {
                'required_columns': ['rough_id', 'weight', 'pieces'],
                'numeric_columns': ['weight', 'pieces', 'purchase_price'],
                'non_empty_columns': ['rough_id']
            },
            'sales.xlsx': {
                'required_columns': ['sale_id', 'date', 'item_id', 'total_amount'],
                'numeric_columns': ['total_amount'],
                'non_empty_columns': ['sale_id', 'item_id']
            },
            'purchases.xlsx': {
                'required_columns': ['purchase_id', 'date', 'total_amount'],
                'numeric_columns': ['total_amount'],
                'non_empty_columns': ['purchase_id']
            }
        }
        
        issues = []
        
        for file_name, rules in validation_rules.items():
            file_path = os.path.join(data_dir, file_name)
            if not os.path.exists(file_path):
                issues.append(f"Missing file: {file_name}")
                continue
                
            try:
                df = pd.read_excel(file_path)
                
                # Check required columns
                missing_columns = [col for col in rules['required_columns'] 
                                 if col not in df.columns]
                if missing_columns:
                    issues.append(f"{file_name}: Missing required columns: {', '.join(missing_columns)}")
                
                # Check numeric columns
                for col in rules['numeric_columns']:
                    if col in df.columns:
                        non_numeric = df[col].apply(lambda x: not pd.api.types.is_numeric_dtype(type(x)))
                        if non_numeric.any():
                            issues.append(f"{file_name}: Non-numeric values in column {col}")
                
                # Check non-empty columns
                for col in rules['non_empty_columns']:
                    if col in df.columns:
                        empty_rows = df[df[col].isna() | (df[col] == '')].index.tolist()
                        if empty_rows:
                            issues.append(f"{file_name}: Empty values in column {col} at rows {empty_rows}")
                
            except Exception as e:
                issues.append(f"Error validating {file_name}: {str(e)}")
        
        if issues:
            logger.warning("Data validation issues found:")
            for issue in issues:
                logger.warning(f"  - {issue}")
            return False
        
        logger.info("Data validation completed successfully")
        return True
    except Exception as e:
        logger.error(f"Error during data validation: {str(e)}")
        return False

def enhance_excel_formatting(file_path, sheet_name='Sheet1'):
    """
    Enhance Excel file formatting with consistent styling.
    """
    try:
        # Check if file exists first
        if not os.path.exists(file_path):
            # Create a new workbook if file doesn't exist
            wb = Workbook()
            sheet = wb.active
            sheet.title = sheet_name
            wb.save(file_path)
            logger.info(f"Created new Excel file at {file_path}")
        
        # Load the workbook directly with openpyxl
        wb = load_workbook(file_path)
        
        # Get the active sheet
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.active
            if sheet is not None:
                sheet.title = sheet_name
        
        # Ensure sheet is not None before proceeding
        if sheet is None:
            raise ValueError(f"Could not find or create sheet '{sheet_name}' in workbook")
        
        # Format headers
        if sheet.max_row > 0:
            for col_num, column_title in enumerate(sheet[1], 1):
                if column_title is None or column_title.value is None:
                    continue
                
                cell = sheet.cell(row=1, column=col_num)
                cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                
                # Adjust column width based on content
                column_letter = get_column_letter(col_num)
                max_length = 0
                for cell in sheet[column_letter]:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = min(adjusted_width, 30)
        
            # Format data cells
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    cell.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
                    
                    # Align numeric values to the right
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.number_format = '#,##0.00'
                    else:
                        cell.alignment = Alignment(vertical='center')
        
        # Save the workbook
        wb.save(file_path)
        logger.info(f"Enhanced formatting for {os.path.basename(file_path)}")
        return True
    except Exception as e:
        logger.error(f"Error enhancing Excel formatting: {str(e)}")
        return False 